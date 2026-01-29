#!/usr/bin/env python3

# wc_person_to_mysql.py
#
# - Default: auto READ CRHD semua WC aktif (LVORM!='X')
# - Filter: --werks-filter (ulang), --like "pattern%" (ulang; %/_ seperti SQL LIKE)
# - Pasangan spesifik: --pairs "ARBPL:WERKS,...", atau --arbpl ... --werks ...
# - Blacklist PERNR: hardcoded + ENV + file; opsi purge global
# - Log: console + rotating harian + unique per-run; path fleksibel via ENV WC_LOG_DIR
# - DB: selective delete per pasangan; collation utf8mb4_unicode_ci
#
# - Tambahan:
#   - cek role NIK via Z_RFC_DISPLAY_NIK_CONF (PERNR saja) dan update kolom `role`
#   - ambil deskripsi WC via Z_FM_GET_WC_DESC (IV_ARBPL/IV_WERKS) dan update kolom `desc`
#   - ambil DEVISI dari Excel (PLANT kolom B, KODE kolom E, DEVISI kolom C) dan update kolom `devisi`
#   - jika dijalankan pada tanggal 6, lakukan backup isi wc_person_data ke tabel backup per bulan

"""
Cara pakai ringkas

1) Mode default: scan CRHD semua WERKS aktif (LVORM != 'X')

   python wc_person_to_mysql.py

2) Batasi plant & pola ARBPL saat auto READ

   python wc_person_to_mysql.py --werks-filter 1000 --werks-filter 2000 --like "WC%"

3) Hanya pasangan spesifik (tanpa tanggal; otomatis 31.12.9999)

   python wc_person_to_mysql.py --pairs "WC034:1000,WC035:2000"

4) Kombinasi ARBPL/WERKS via argumen terpisah

   python wc_person_to_mysql.py --arbpl WC034 --arbpl WC035 --werks 1000 --werks 2000

5) Lihat rencana saja (tanpa ubah DB) + tampilkan daftar pasangan

   python wc_person_to_mysql.py --dry-run --show-pairs --verbose-steps

6) Blacklist PERNR (skip insert) + purge data lama
   # tambahan blacklist dari file TXT/CSV (kolom PERNR) dan hapus dari DB:

   python wc_person_to_mysql.py --blacklist-file my_blacklist.csv --purge-blacklist

   # atau dari ENV (comma separated):

   set WC_BLACKLIST_PERNR=10001234,10005678

   python wc_person_to_mysql.py
"""

import os
import sys
import re
import csv
import argparse
import signal
import time
import logging
import datetime
from typing import Any, Dict, List, Optional, Tuple, Set
from pathlib import Path
from logging.handlers import TimedRotatingFileHandler

from dotenv import load_dotenv
from pyrfc import (
    Connection,
    CommunicationError,
    LogonError,
    ABAPApplicationError,
    ABAPRuntimeError,
)
import mysql.connector

# untuk baca Excel DEVISI
try:
    from openpyxl import load_workbook  # pip install openpyxl
except ImportError:
    load_workbook = None

# --- pyrfc lama (refer 'long' Python2)
import builtins as _bt

if not hasattr(_bt, "long"):
    _bt.long = int

try:
    signal.signal(signal.SIGINT, signal.SIG_IGN)
except Exception:
    pass

load_dotenv()

# ---------- SAP ----------
DEFAULT_SAP = {
    "ashost": os.environ.get("SAP_ASHOST", "192.168.254.154"),
    "sysnr": os.environ.get("SAP_SYSNR", "01"),
    "client": os.environ.get("SAP_CLIENT", "300"),
    "lang": os.environ.get("SAP_LANG", "EN"),
}

SAP_USERNAME = os.environ.get("SAP_USER", "auto_email")
SAP_PASSWORD = os.environ.get("SAP_PASS", "11223344")

RFC_NAME = "CR_PERSONS_OF_WORKCENTER"
RFC_ROLE_NAME = "Z_RFC_DISPLAY_NIK_CONF"  # RFC kedua untuk cek role NIK (PERNR saja)
RFC_DESC_NAME = "Z_FM_GET_WC_DESC"        # RFC ketiga untuk ambil deskripsi WC (ARBPL/WERKS)

# ---------- MySQL ----------
DB_HOST = os.environ.get("DB_HOST", "127.0.0.1")
DB_PORT = int(os.environ.get("DB_PORT", "3306"))
DB_USER = os.environ.get("DB_USER", "root")
DB_PASS = os.environ.get("DB_PASS", "")
DB_NAME = os.environ.get("DB_NAME", "wc_person")
TABLE = os.environ.get("DB_TABLE", "wc_person_data")
BACKUP_TABLE = os.environ.get("DB_BACKUP_TABLE", "wc_person_backup")  # tabel backup per bulan

# ---------- Logging ----------
PROJECT_ROOT = Path(__file__).resolve().parent

DEFAULT_LOG_DIR = PROJECT_ROOT / "storage" / "logs" / "python_wc_person_mysql"
LOG_DIR = Path(os.environ.get("WC_LOG_DIR", str(DEFAULT_LOG_DIR)))
LOG_DIR.mkdir(parents=True, exist_ok=True)

_start_ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
RUN_LOG_FILE = LOG_DIR / f"wc_person_to_mysql_{os.getpid()}_{_start_ts}.log"
MAIN_LOG_FILE = LOG_DIR / "wc_person_to_mysql.log"

# lokasi file Excel DEVISI (bisa override via ENV WC_DEVISI_FILE)
DEVISI_FILE = Path(os.environ.get("WC_DEVISI_FILE", str(PROJECT_ROOT / "DEVISI.xlsx")))

logger = logging.getLogger("wc_person")
logger.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s | %(message)s", datefmt="%H:%M:%S")

# Console
sh = logging.StreamHandler(sys.stdout)
sh.setLevel(logging.INFO)
sh.setFormatter(formatter)
logger.addHandler(sh)

# Rotating harian
rot = TimedRotatingFileHandler(
    str(MAIN_LOG_FILE),
    when="midnight",
    interval=1,
    backupCount=15,
    encoding="utf-8",
    utc=False,
)
rot.setLevel(logging.INFO)
rot.setFormatter(formatter)
logger.addHandler(rot)

# File unik per-run
fh = logging.FileHandler(str(RUN_LOG_FILE), mode="w", encoding="utf-8")
fh.setLevel(logging.INFO)
fh.setFormatter(formatter)
logger.addHandler(fh)

# ---------- Util ----------
DEFAULT_DATE_STR = "31.12.9999"
WIDTH = 80


def hr(ch: str = "=") -> None:
    logger.info(ch * WIDTH)


def subhr() -> None:
    logger.info("-" * WIDTH)


def title(s: str) -> None:
    hr("=")
    logger.info(s)
    hr("=")


def to_dats(s: str) -> str:
    s = s.strip()
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", s)  # DD.MM.YYYY
    if m:
        dd, mm, yyyy = m.groups()
        return f"{yyyy}{mm}{dd}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)  # YYYY-MM-DD
    if m:
        yyyy, mm, dd = m.groups()
        return f"{yyyy}{mm}{dd}"
    m = re.match(r"^\d{8}$", s)  # YYYYMMDD
    if m:
        return s
    raise ValueError(f"Format tanggal tidak dikenali: {s}")


def parse_pairs_args(
    arbpls: List[str], werks_list: List[str], dats_const: str
) -> List[Tuple[str, str, str]]:
    out: List[Tuple[str, str, str]] = []
    if not arbpls and not werks_list:
        return out
    arbpls = arbpls or ["WC034"]
    werks_list = werks_list or ["1000"]
    for a in arbpls:
        for w in werks_list:
            out.append((a.strip(), w.strip(), dats_const))
    return out


def parse_pairs_string(pairs: str, dats_const: str) -> List[Tuple[str, str, str]]:
    if not pairs:
        return []
    out: List[Tuple[str, str, str]] = []
    for chunk in pairs.split(","):
        parts = [p.strip() for p in chunk.split(":")]
        if len(parts) == 2:
            a, w = parts
        elif len(parts) == 3:
            a, w, _ = parts
        else:
            raise ValueError(f"Format pairs tidak valid: {chunk}")
        out.append((a, w, dats_const))
    return out


def parse_pairs_csv(path: str, dats_const: str) -> List[Tuple[str, str, str]]:
    if not path:
        return []
    rows: List[Tuple[str, str, str]] = []
    with open(path, newline="", encoding="utf-8") as f:
        rdr = csv.DictReader(f)
        need = {"ARBPL", "WERKS"}
        hdrs = {h.upper(): h for h in (rdr.fieldnames or [])}
        if not need.issubset(set(hdrs.keys())):
            raise ValueError("Header CSV harus minimal: ARBPL,WERKS")
        for r in rdr:
            rows.append(
                (r[hdrs["ARBPL"]].strip(), r[hdrs["WERKS"]].strip(), dats_const)
            )
    return rows


def dedupe_pairs(pairs: List[Tuple[str, str, str]]) -> List[Tuple[str, str, str]]:
    seen: Set[Tuple[str, str, str]] = set()
    out: List[Tuple[str, str, str]] = []
    for a, w, d in pairs:
        key = (a.strip().upper(), w.strip().upper(), d)
        if key in seen:
            continue
        seen.add(key)
        out.append((a.strip(), w.strip(), d))
    return out


def like_to_regex(pattern: str) -> re.Pattern:
    esc = re.escape(pattern)
    esc = esc.replace(r"\%", ".*").replace(r"\_", ".")
    return re.compile(f"^{esc}$", re.IGNORECASE)


def rows_from_rfc(conn: Connection, arbpl: str, werks: str, dats: str) -> List[Dict[str, Any]]:
    resp = conn.call(RFC_NAME, ARBPL=arbpl, WERKS=werks, DATE=dats)
    for _k, v in resp.items():
        if isinstance(v, list) and (not v or isinstance(v[0], dict)):
            return v
    return []


def normalize_row(
    r: Dict[str, Any], default_arbpl: str, default_werks: str
) -> Dict[str, Any]:
    def S(x: Any) -> str:
        return "" if x is None else str(x)

    arbpl_val = S(r.get("ARBPL"))
    werks_val = S(r.get("WERKS"))
    return {
        "otype": S(r.get("OTYPE")),
        "objid": S(r.get("OBJID")),
        "pernr": S(r.get("PERNR")).zfill(8)
        if S(r.get("PERNR")).isdigit()
        else S(r.get("PERNR")),
        "begda": S(r.get("BEGDA")),
        "endda": S(r.get("ENDDA")),
        "arbid": S(r.get("ARBID")),
        "short": S(r.get("SHORT")),
        "stext": S(r.get("STEXT")),
        "arbpl": arbpl_val if arbpl_val else default_arbpl,
        "werks": werks_val if werks_val else default_werks,
    }

# ---------- Blacklist PERNR ----------

HARDCODED_BLACKLIST = {
    "10000011",
    "10000015",
    "10000040",
    "10000062",
    "10000063",
    "10000083",
    "10000110",
    "10000126",
    "10000144",
    "10000161",
    "10000189",
    "10000364",
    "10000395",
    "10000414",
    "10000417",
    "10000427",
    "10000431",
    "10000440",
    "10000458",
    "10000482",
    "10000502",
    "10000524",
    "10000541",
    "10000548",
    "10000555",
    "10000564",
    "10000570",
    "10000577",
    "10000591",
    "10000615",
    "10000622",
    "10000642",
    "10000659",
    "10000725",
    "10000778",
    "10000874",
    "10001561",
    "10001983",
    "10002308",
    "10002690",
    "10002787",
    "10003007",
    "10003008",
    "10003009",
    "10004934",
    "10004994",
    "10005063",
    "10003590",
    "10003599",
    "10003600",
    "10004874",
    "10000897",
    "10000898",
    "10001002",
    "10006163",
    "10006337",
    "10007161",
    "10002446",
    "10000467",
    "10004411",
    "10000644",
    "10000026",
    "10000093",
    "10000109",
    "10000112",
    "10000141",
    "10000266",
    "10000319",
    "10002804",
    "10000420",
    "10005689",
    "10008126",
    "10008135",
    "10008134",
}


def _pad8(s: str) -> str:
    s = (s or "").strip()
    return s.zfill(8) if s.isdigit() else s


def load_blacklist(extra_file: str) -> set:
    bl = {_pad8(x) for x in HARDCODED_BLACKLIST}
    env_val = os.environ.get("WC_BLACKLIST_PERNR", "")
    for tok in [t.strip() for t in env_val.split(",") if t.strip()]:
        bl.add(_pad8(tok))

    if extra_file:
        try:
            with open(extra_file, "r", encoding="utf-8") as f:
                head = f.read(2048)
                f.seek(0)
                if "," in head and "pernr" in head.lower():
                    rdr = csv.DictReader(f)
                    for r in rdr:
                        for k in r.keys():
                            if k and k.lower() == "pernr":
                                bl.add(_pad8(str(r[k])))
                else:
                    for line in f:
                        for tok in re.split(r"[\s,;]+", line.strip()):
                            if tok:
                                bl.add(_pad8(tok))
        except Exception as e:
            logger.info(f"[WARN] Gagal baca blacklist-file: {e}")
    return bl

# ---------- RFC_READ_TABLE (CRHD) ----------

def fetch_wc_pairs_from_crhd(
    conn: Connection,
    valid_werks: Optional[List[str]] = None,
    like_patterns: Optional[List[str]] = None,
) -> List[Tuple[str, str]]:
    try:
        res = conn.call(
            "RFC_READ_TABLE",
            QUERY_TABLE="CRHD",
            DELIMITER="|",
            FIELDS=[
                {"FIELDNAME": "ARBPL"},
                {"FIELDNAME": "WERKS"},
                {"FIELDNAME": "LVORM"},
            ],
            ROWCOUNT=100000,
        )
    except Exception as e:
        logger.info(f"[WARN] RFC_READ_TABLE CRHD gagal: {e}")
        return []

    like_regexes = [like_to_regex(p) for p in (like_patterns or [])]
    pairs_set: Set[Tuple[str, str]] = set()

    for row in res.get("DATA", []):
        try:
            arbpl, werks, lvorm = row["WA"].split("|")
            arbpl = arbpl.strip()
            werks = werks.strip()
            lvorm = lvorm.strip()
        except Exception:
            continue
        if lvorm == "X":
            continue
        if valid_werks and werks not in valid_werks:
            continue
        if like_regexes and not any(rgx.match(arbpl) for rgx in like_regexes):
            continue
        if arbpl and werks:
            pairs_set.add((arbpl, werks))

    return sorted(pairs_set)

# ---------- MySQL helpers & DDL ----------

DDL_DB = f"""
CREATE DATABASE IF NOT EXISTS `{DB_NAME}`
  DEFAULT CHARACTER SET = utf8mb4
  COLLATE = utf8mb4_unicode_ci;
"""

# UBAH DISINI: Tambahkan kode_laravel
DDL_TABLE = f"""
CREATE TABLE IF NOT EXISTS `{TABLE}` (
  `id` BIGINT UNSIGNED NOT NULL AUTO_INCREMENT,
  `otype`  VARCHAR(4)  NULL,
  `objid`  VARCHAR(20) NOT NULL,
  `pernr`  VARCHAR(12) NOT NULL,
  `begda`  CHAR(8)     NOT NULL,
  `endda`  CHAR(8)     NOT NULL,
  `arbid`  VARCHAR(20) NULL,
  `short`  VARCHAR(80) NULL,
  `stext`  VARCHAR(255) NULL,
  `arbpl`  VARCHAR(30) NOT NULL,
  `werks`  VARCHAR(10) NOT NULL,
  `role`   VARCHAR(20) NULL,
  `desc`   VARCHAR(255) NULL,
  `devisi` VARCHAR(100) NULL,
  `kode_laravel` VARCHAR(255) NULL, -- TAMBAHAN BARU
  `source_rfc` VARCHAR(64) NOT NULL DEFAULT 'CR_PERSONS_OF_WORKCENTER',
  `inserted_at` TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (`id`),
  UNIQUE KEY `uniq_wc_person` (`objid`,`pernr`,`begda`,`arbpl`,`werks`),
  KEY `idx_pernr` (`pernr`),
  KEY `idx_arbpl` (`arbpl`),
  KEY `idx_werks` (`werks`)
) ENGINE=InnoDB
  DEFAULT CHARSET=utf8mb4
  COLLATE=utf8mb4_unicode_ci;
"""

# UBAH DISINI: Tambahkan kode_laravel agar backup juga menyimpan datanya
DDL_BACKUP_TABLE = f"""
CREATE TABLE IF NOT EXISTS `{BACKUP_TABLE}` (
  `id` BIGINT UNSIGNED NOT NULL AUTO_INCREMENT,
  `backup_month` CHAR(7) NOT NULL,  -- format: YYYY-MM (bulan yang dibackup = bulan lalu)
  `otype`  VARCHAR(4)  NULL,
  `objid`  VARCHAR(20) NOT NULL,
  `pernr`  VARCHAR(12) NOT NULL,
  `begda`  CHAR(8)     NOT NULL,
  `endda`  CHAR(8)     NOT NULL,
  `arbid`  VARCHAR(20) NULL,
  `short`  VARCHAR(80) NULL,
  `stext`  VARCHAR(255) NULL,
  `arbpl`  VARCHAR(30) NOT NULL,
  `werks`  VARCHAR(10) NOT NULL,
  `role`   VARCHAR(20) NULL,
  `desc`   VARCHAR(255) NULL,
  `devisi` VARCHAR(100) NULL,
  `kode_laravel` VARCHAR(255) NULL, -- TAMBAHAN BARU DI BACKUP
  `source_rfc` VARCHAR(64) NOT NULL DEFAULT 'CR_PERSONS_OF_WORKCENTER',
  `inserted_at` TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (`id`),
  KEY `idx_backup_month` (`backup_month`),
  KEY `idx_pernr_b` (`pernr`),
  KEY `idx_arbpl_b` (`arbpl`),
  KEY `idx_werks_b` (`werks`)
) ENGINE=InnoDB
  DEFAULT CHARSET=utf8mb4
  COLLATE=utf8mb4_unicode_ci;
"""

# TIDAK ADA PERUBAHAN DI SINI
UPSERT_SQL = f"""
INSERT INTO `{TABLE}`
(`otype`,`objid`,`pernr`,`begda`,`endda`,`arbid`,`short`,`stext`,`arbpl`,`werks`,`source_rfc`)
VALUES (%(otype)s,%(objid)s,%(pernr)s,%(begda)s,%(endda)s,%(arbid)s,%(short)s,%(stext)s,%(arbpl)s,%(werks)s,'CR_PERSONS_OF_WORKCENTER')
ON DUPLICATE KEY UPDATE
  `endda`=VALUES(`endda`),
  `short`=VALUES(`short`),
  `stext`=VALUES(`stext`),
  `arbpl`=VALUES(`arbpl`),
  `werks`=VALUES(`werks`),
  `inserted_at`=CURRENT_TIMESTAMP
"""

DELETE_SQL = f"DELETE FROM `{TABLE}` WHERE `arbpl`=%s AND `werks`=%s"
COUNT_SQL = f"SELECT COUNT(*) FROM `{TABLE}` WHERE `arbpl`=%s AND `werks`=%s"


def get_mysql_conn(database: Optional[str] = None):
    return mysql.connector.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASS,
        database=database,
        autocommit=False,
    )


def ensure_db_and_table():
    # 1) Pastikan DATABASE ada
    root = get_mysql_conn(None)
    cur = root.cursor()
    cur.execute(DDL_DB)
    root.commit()
    cur.close()
    root.close()

    # 2) Konek ke DB yang sudah dipastikan ada
    conn = get_mysql_conn(DB_NAME)
    cur2 = conn.cursor()

    # 3) Pastikan tabel utama & tabel backup ada (pakai DDL terbaru)
    cur2.execute(DDL_TABLE)
    cur2.execute(DDL_BACKUP_TABLE)

    # ------------------------ MIGRASI TABEL UTAMA ------------------------ #
    # Tambah kolom role kalau belum ada (untuk instalasi lama)
    cur2.execute(f"SHOW COLUMNS FROM `{TABLE}` LIKE 'role'")
    if not cur2.fetchone():
        cur2.execute(
            f"ALTER TABLE `{TABLE}` "
            "ADD COLUMN `role` VARCHAR(20) NULL AFTER `werks`"
        )
        logger.info(f"[DB] Kolom 'role' ditambahkan ke {TABLE}.")

    # Tambah kolom desc (deskripsi WC) kalau belum ada
    cur2.execute(f"SHOW COLUMNS FROM `{TABLE}` LIKE 'desc'")
    if not cur2.fetchone():
        cur2.execute(
            f"ALTER TABLE `{TABLE}` "
            "ADD COLUMN `desc` VARCHAR(255) NULL AFTER `role`"
        )
        logger.info(f"[DB] Kolom 'desc' ditambahkan ke {TABLE}.")

    # Tambah kolom devisi kalau belum ada
    cur2.execute(f"SHOW COLUMNS FROM `{TABLE}` LIKE 'devisi'")
    if not cur2.fetchone():
        cur2.execute(
            f"ALTER TABLE `{TABLE}` "
            "ADD COLUMN `devisi` VARCHAR(100) NULL AFTER `desc`"
        )
        logger.info(f"[DB] Kolom 'devisi' ditambahkan ke {TABLE}.")

    # Tambah kolom kode_laravel kalau belum ada
    cur2.execute(f"SHOW COLUMNS FROM `{TABLE}` LIKE 'kode_laravel'")
    if not cur2.fetchone():
        cur2.execute(
            f"ALTER TABLE `{TABLE}` "
            "ADD COLUMN `kode_laravel` VARCHAR(255) NULL AFTER `devisi`"
        )
        logger.info(f"[DB] Kolom 'kode_laravel' ditambahkan ke {TABLE}.")

    # ------------------------ MIGRASI TABEL BACKUP ----------------------- #
    # Pastikan kolom-kolom baru juga ada di tabel BACKUP (untuk instalasi lama)
    cur2.execute(f"SHOW COLUMNS FROM `{BACKUP_TABLE}` LIKE 'role'")
    if not cur2.fetchone():
        cur2.execute(
            f"ALTER TABLE `{BACKUP_TABLE}` "
            "ADD COLUMN `role` VARCHAR(20) NULL AFTER `werks`"
        )
        logger.info(f"[DB] Kolom 'role' ditambahkan ke {BACKUP_TABLE}.")

    cur2.execute(f"SHOW COLUMNS FROM `{BACKUP_TABLE}` LIKE 'desc'")
    if not cur2.fetchone():
        cur2.execute(
            f"ALTER TABLE `{BACKUP_TABLE}` "
            "ADD COLUMN `desc` VARCHAR(255) NULL AFTER `role`"
        )
        logger.info(f"[DB] Kolom 'desc' ditambahkan ke {BACKUP_TABLE}.")

    cur2.execute(f"SHOW COLUMNS FROM `{BACKUP_TABLE}` LIKE 'devisi'")
    if not cur2.fetchone():
        cur2.execute(
            f"ALTER TABLE `{BACKUP_TABLE}` "
            "ADD COLUMN `devisi` VARCHAR(100) NULL AFTER `desc`"
        )
        logger.info(f"[DB] Kolom 'devisi' ditambahkan ke {BACKUP_TABLE}.")

    cur2.execute(f"SHOW COLUMNS FROM `{BACKUP_TABLE}` LIKE 'kode_laravel'")
    if not cur2.fetchone():
        cur2.execute(
            f"ALTER TABLE `{BACKUP_TABLE}` "
            "ADD COLUMN `kode_laravel` VARCHAR(255) NULL AFTER `devisi`"
        )
        logger.info(f"[DB] Kolom 'kode_laravel' ditambahkan ke {BACKUP_TABLE}.")

    # 4) Selesai
    cur2.close()
    conn.commit()
    return conn
# ---------- Role checker via Z_RFC_DISPLAY_NIK_CONF ----------

def check_and_update_roles(
    rfc: Connection,
    db,
    role_pernrs: Set[str],
    role_value: str = "INDUK",
) -> None:
    """
    Cek setiap PERNR ke RFC Z_RFC_DISPLAY_NIK_CONF.
    - Panggil: Z_RFC_DISPLAY_NIK_CONF(PERNR=<pernr>)
    - Jika RESULT_TAB ada baris aktif (DELETED != 'X') -> role=role_value
    Update semua baris di tabel MySQL dengan PERNR tersebut.
    """
    if not role_pernrs:
        return

    title(f"CEK ROLE NIK via {RFC_ROLE_NAME}")

    active_pernrs: Set[str] = set()
    pernrs_sorted = sorted(role_pernrs)
    total = len(pernrs_sorted)

    for idx, pernr in enumerate(pernrs_sorted, start=1):
        if idx == 1 or idx % 50 == 0:
            logger.info(f"  Progress role {idx}/{total} ... (PERNR={pernr})")

        try:
            # Hanya kirim PERNR; WERKS optional & tidak dikirim
            resp = rfc.call(RFC_ROLE_NAME, PERNR=pernr)
        except (ABAPApplicationError, ABAPRuntimeError, CommunicationError) as e:
            logger.info(f"  [WARN] RFC {RFC_ROLE_NAME} PERNR={pernr}: {e}")
            continue

        tab = resp.get("RESULT_TAB") or []
        is_active = any(
            str(row.get("DELETED", "")).strip().upper() != "X" for row in tab
        )
        if is_active:
            active_pernrs.add(pernr)

    if not active_pernrs:
        logger.info(
            f"Tidak ada NIK yang berstatus role='{role_value}' dari RFC {RFC_ROLE_NAME}."
        )
        return

    cur = db.cursor()
    sql = f"UPDATE `{TABLE}` SET `role`=%s WHERE `pernr`=%s"
    data = [(role_value, p) for p in active_pernrs]

    try:
        cur.executemany(sql, data)
        db.commit()
        logger.info(f"[ROLE] Ter-update {cur.rowcount} baris role='{role_value}'.")
    except mysql.connector.Error as e:
        db.rollback()
        logger.info(f"[ERROR] Update role di DB: {e}")
    finally:
        cur.close()

# ---------- WC description via Z_FM_GET_WC_DESC ----------

def update_wc_descriptions(
    rfc: Connection,
    db,
    pairs: List[Tuple[str, str, str]],
) -> None:
    """
    Ambil deskripsi WC (E_DESC) dari Z_FM_GET_WC_DESC untuk setiap pasangan
    ARBPL/WERKS yang terlibat, lalu update kolom `desc` di tabel MySQL.
    """
    # Ambil pasangan unik ARBPL/WERKS saja
    unique_pairs = sorted({(a, w) for (a, w, _d) in pairs})
    if not unique_pairs:
        return

    title(f"UPDATE DESKRIPSI WC via {RFC_DESC_NAME}")

    cur = db.cursor()
    sql = f"UPDATE `{TABLE}` SET `desc`=%s WHERE `arbpl`=%s AND `werks`=%s"

    total_updated = 0
    total_pairs = len(unique_pairs)

    for idx, (arbpl, werks) in enumerate(unique_pairs, start=1):
        if idx == 1 or idx % 50 == 0:
            logger.info(
                f"  Progress desc {idx}/{total_pairs} ... "
                f"(ARBPL={arbpl}, WERKS={werks})"
            )

        try:
            # Panggil RFC: kirim ARBPL + WERKS
            resp = rfc.call(RFC_DESC_NAME, IV_ARBPL=arbpl, IV_WERKS=werks)
        except (ABAPApplicationError, ABAPRuntimeError, CommunicationError) as e:
            logger.info(
                f"  [WARN] RFC {RFC_DESC_NAME} ARBPL={arbpl} WERKS={werks}: {e}"
            )
            continue

        desc_val = (resp.get("E_DESC") or "").strip()

        # Kalau kosong, biarkan nilai lama apa adanya
        if not desc_val:
            continue

        try:
            cur.execute(sql, (desc_val, arbpl, werks))
            db.commit()
            total_updated += cur.rowcount
        except mysql.connector.Error as e:
            db.rollback()
            logger.info(
                f"  [ERROR] Update desc untuk ARBPL={arbpl}, WERKS={werks}: {e}"
            )

    cur.close()
    logger.info(f"[DESC] Ter-update {total_updated} baris deskripsi WC.")

# ---------- DEVISI mapping via Excel ----------

def load_devisi_mapping_from_excel(path: Path) -> Dict[Tuple[str, str], Dict[str, str]]:
    """
    Baca Excel:

    - Kolom B (idx 1): PLANT
    - Kolom C (idx 2): DEVISI
    - Kolom E (idx 4): KODE (ARBPL)
    - Kolom F (idx 5): KODE LARAVEL (bisa merged ke beberapa baris)
    """

    mapping: Dict[Tuple[str, str], Dict[str, str]] = {}

    if load_workbook is None:
        logger.info("[EXCEL] Modul openpyxl tidak tersedia.")
        return mapping

    p = Path(path)
    if not p.is_file():
        logger.info(f"[EXCEL] File tidak ditemukan: {p}")
        return mapping

    try:
        wb = load_workbook(filename=str(p), data_only=True)
        ws = wb.active
    except Exception as e:
        logger.info(f"[EXCEL] Gagal buka file: {e}")
        return mapping

    # --- BACA MERGED CELL UNTUK KOLOM F (KODE LARAVEL) ---
    # mapping: row_index -> nilai kode_laravel hasil merge
    merged_kolar_by_row: Dict[int, str] = {}
    for cell_range in ws.merged_cells.ranges:
        # Kita hanya peduli range yang menyentuh kolom F (kolom ke-6)
        if cell_range.min_col <= 6 <= cell_range.max_col:
            # Ambil nilai dari sel paling atas di kolom F pada range tersebut
            top_cell = ws.cell(row=cell_range.min_row, column=6)
            if top_cell.value not in (None, ""):
                val = str(top_cell.value).strip()
                for r in range(cell_range.min_row, cell_range.max_row + 1):
                    merged_kolar_by_row[r] = val

    current_devisi = ""
    row_mapped = 0

    # Asumsi baris 1 header, data mulai baris 2
    for row in ws.iter_rows(min_row=2):
        # Ambil Cell
        plant_cell = row[1] if len(row) > 1 else None  # B
        devisi_cell = row[2] if len(row) > 2 else None  # C
        kode_cell = row[4] if len(row) > 4 else None    # E
        kolar_cell = row[5] if len(row) > 5 else None   # F

        plant_val = ""
        kode_val = ""

        # PLANT (B)
        if plant_cell and plant_cell.value is not None:
            plant_val = str(plant_cell.value).strip()

        # DEVISI (C) – tetap pakai mekanisme "last non-empty"
        if devisi_cell and devisi_cell.value not in (None, ""):
            current_devisi = str(devisi_cell.value).strip()

        # KODE / ARBPL (E)
        if kode_cell and kode_cell.value is not None:
            kode_val = str(kode_cell.value).strip()

        # KODE LARAVEL (F)
        kode_laravel_val = ""
        if kolar_cell:
            # 1) Kalau di sel F ini ada nilai → pakai langsung
            if kolar_cell.value not in (None, ""):
                kode_laravel_val = str(kolar_cell.value).strip()
            else:
                # 2) Sel kosong, cek apakah baris ini bagian dari merged cell di kolom F
                merged_val = merged_kolar_by_row.get(kolar_cell.row)
                if merged_val is not None:
                    kode_laravel_val = merged_val
                # 3) Kalau tidak ada di merged_kolar_by_row → biarkan kosong ("")
        # Kalau plant atau kode kosong, skip
        if not plant_val or not kode_val:
            continue

        # Simpan ke mapping. Key=(ARBPL, WERKS/PLANT)
        key = (kode_val.upper(), plant_val)
        if key not in mapping:
            mapping[key] = {
                "devisi": current_devisi,
                "kode_laravel": kode_laravel_val,
            }
            row_mapped += 1

    logger.info(f"[EXCEL] Total mapping loaded: {row_mapped}")
    return mapping



def update_devisi_from_excel(
    db,
    pairs: List[Tuple[str, str, str]],
    mapping: Dict[Tuple[str, str], Dict[str, str]], # Type hint berubah
) -> None:
    
    if not mapping:
        return

    # Set pasangan yang ikut di run sekarang
    run_pairs = {(a.strip().upper(), w.strip()) for (a, w, _d) in pairs}

    # Susun list update: (devisi, kode_laravel, arbpl, werks)
    # Susun list update: (devisi, kode_laravel, arbpl, werks)
    updates: List[Tuple[str, str, str, str]] = []
    
    for (arbpl, werks), data in mapping.items():
        key = (arbpl.strip().upper(), werks.strip())
        
        # Cek apakah pasangan ini ada di list proses saat ini
        if key in run_pairs:
            # Ambil nilai DEVISI apa adanya (boleh kosong)
            dev = data.get("devisi", "")

            # Ambil kode_laravel, kalau kosong => None (NULL di MySQL)
            raw_kolar = (data.get("kode_laravel") or "").strip()
            kolar = raw_kolar if raw_kolar else None   # <- INI YANG PENTING

            updates.append((dev, kolar, arbpl, werks))

    if not updates:
        logger.info("[EXCEL] Tidak ada data Excel yang cocok dengan run ini.")
        return

    title("UPDATE INFO DARI EXCEL (DEVISI & KODE LARAVEL)")
    
    cur = db.cursor()
    # QUERY UPDATE DITAMBAH: kode_laravel
    sql = """
        UPDATE `{TABLE}` 
        SET `devisi`=%s, `kode_laravel`=%s 
        WHERE `arbpl`=%s AND `werks`=%s
    """.replace("{TABLE}", TABLE) 

    total_updated = 0
    batch_size = 200

    try:
        for i in range(0, len(updates), batch_size):
            chunk = updates[i:i + batch_size]
            cur.executemany(sql, chunk)
            total_updated += cur.rowcount
        db.commit()
        logger.info(f"[EXCEL] Berhasil update {total_updated} baris.")
    except mysql.connector.Error as e:
        db.rollback()
        logger.info(f"[EXCEL] ERROR Update DB: {e}")
    finally:
        cur.close()

# ---------- Backup bulanan ----------

def _previous_month(today: datetime.date) -> Tuple[int, int]:
    """
    Mengembalikan (tahun, bulan) untuk bulan sebelumnya.
    Contoh: 2025-03-03 -> (2025, 2)
            2025-01-03 -> (2024, 12)
    """
    year = today.year
    month = today.month
    if month == 1:
        return year - 1, 12
    return year, month - 1


def do_monthly_backup_if_needed(db) -> None:
    """
    Backup data wc_person_data ke tabel backup HANYA jika hari ini tanggal 6.
    - Backup disimpan per bulan (kolom backup_month = 'YYYY-MM' bulan LALU)
    - Setiap tanggal 6, backup untuk bulan itu dihapus lalu di-insert ulang (refresh).
    """
    today = datetime.date.today()
    logger.info(f"[BACKUP] Tanggal hari ini: {today.isoformat()}")

    # Hanya jalan di tanggal 6
    if today.day != 6:
        logger.info("[BACKUP] Bukan tanggal 6, backup di-skip.")
        return

    # Hitung bulan sebelumnya
    year, month = _previous_month(today)
    backup_month_str = f"{year:04d}-{month:02d}"
    logger.info(f"[BACKUP] Proses backup untuk bulan: {backup_month_str}")

    cur = db.cursor()
    try:
        # 1) Hapus backup lama untuk bulan tsb
        delete_sql = f"DELETE FROM `{BACKUP_TABLE}` WHERE `backup_month` = %s"
        cur.execute(delete_sql, (backup_month_str,))
        logger.info(f"[BACKUP] Hapus backup lama: {cur.rowcount} baris.")

        # 2) Insert dari tabel utama ke tabel backup
        insert_sql = f"""
            INSERT INTO `{BACKUP_TABLE}` (
                backup_month,
                otype, objid, pernr, begda, endda,
                arbid, short, stext, arbpl, werks,
                role, `desc`, devisi, kode_laravel,
                source_rfc, inserted_at
            )
            SELECT
                %s AS backup_month,
                otype, objid, pernr, begda, endda,
                arbid, short, stext, arbpl, werks,
                role, `desc`, devisi, kode_laravel,
                source_rfc, inserted_at
            FROM `{TABLE}`
        """
        cur.execute(insert_sql, (backup_month_str,))
        logger.info(f"[BACKUP] Baris yang dibackup: {cur.rowcount}")

        db.commit()
        logger.info("[BACKUP] Backup bulanan selesai dan di-commit.")
    except mysql.connector.Error as e:
        db.rollback()
        logger.info(f"[BACKUP] ERROR backup bulanan: {e}")
    finally:
        cur.close()

# ---------- Main ----------

def main():
    ap = argparse.ArgumentParser(
        description="Load CR_PERSONS_OF_WORKCENTER (OUT_PERSONS) ke MySQL."
    )

    # Input manual
    ap.add_argument("--arbpl", action="append", help="ARBPL; bisa diulang.")
    ap.add_argument("--werks", action="append", help="WERKS; bisa diulang.")
    ap.add_argument("--pairs", default="", help='Daftar "ARBPL:WERKS,ARBPL2:WERKS2"')
    ap.add_argument("--pairs-file", default="", help="CSV 2 kolom: ARBPL,WERKS")

    # Mode auto READ filter
    ap.add_argument("--werks-filter", action="append", help="Batasi plant auto READ")
    ap.add_argument("--like", action="append", help='Filter ARBPL ala LIKE "WC%"')

    # Global date override
    ap.add_argument(
        "--date-all",
        default=DEFAULT_DATE_STR,
        help="Override tanggal global (default: 31.12.9999)",
    )

    # Blacklist
    ap.add_argument(
        "--blacklist-file",
        default="",
        help="TXT/CSV kolom PERNR tambahan blacklist",
    )
    ap.add_argument(
        "--purge-blacklist",
        action="store_true",
        help="Hapus baris di DB untuk semua PERNR blacklist (global)",
    )

    # Log & mode
    ap.add_argument("--show-pairs", action="store_true", help="Tampilkan pasangan")
    ap.add_argument(
        "--verbose-steps",
        action="store_true",
        help="Log detail per langkah (RFC/DB)",
    )
    ap.add_argument(
        "--no-delete",
        action="store_true",
        help="Jangan hapus data lama untuk pasangan ini",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Simulasi: tidak ubah DB, hanya log",
    )
    ap.add_argument(
        "--batch",
        type=int,
        default=500,
        help="Ukuran batch insert (default: 500)",
    )

    args = ap.parse_args()
    dats_const = to_dats(args.date_all)

    logger.info(f"LOG DIR       : {LOG_DIR}")
    logger.info(f"MAIN LOG      : {MAIN_LOG_FILE}")
    logger.info(f"RUN LOG       : {RUN_LOG_FILE}")
    logger.info(f"DEVISI FILE   : {DEVISI_FILE}")

    job_start_wall = datetime.datetime.now().astimezone()
    job_t0 = time.perf_counter()

    # SAP connect
    sap_params = {**DEFAULT_SAP, "user": SAP_USERNAME, "passwd": SAP_PASSWORD}
    title(
        f"Connect SAP {sap_params['ashost']} client {sap_params['client']} as {SAP_USERNAME} ..."
    )
    try:
        rfc = Connection(**sap_params)
    except (CommunicationError, LogonError) as e:
        logger.info(f"Gagal konek SAP: {e}")
        logger.info(f"Log tersimpan di: {RUN_LOG_FILE}")
        sys.exit(2)
    logger.info(
        f"OK. DATE efektif untuk semua panggilan: {args.date_all} -> {dats_const}\n"
    )

    # Blacklist
    blacklist = load_blacklist(args.blacklist_file)
    if blacklist:
        logger.info(f"Blacklist aktif: {len(blacklist)} PERNR")

    # Susun pasangan dari argumen
    pairs: List[Tuple[str, str, str]] = []
    pairs += parse_pairs_args(args.arbpl or [], args.werks or [], dats_const)
    pairs += parse_pairs_string(args.pairs, dats_const)
    pairs += parse_pairs_csv(args.pairs_file, dats_const)
    pairs = dedupe_pairs(pairs)

    # Kalau kosong → auto READ CRHD
    if not pairs:
        valid_werks = args.werks_filter or []
        like_patterns = args.like or []
        logger.info(
            "Tidak ada pasangan dari argumen. Auto READ CRHD (RFC_READ_TABLE) ..."
        )
        wc_pairs = fetch_wc_pairs_from_crhd(
            rfc, valid_werks=valid_werks, like_patterns=like_patterns
        )
        logger.info(f"Ditemukan {len(wc_pairs)} pasangan aktif di CRHD.")
        pairs = [(a, w, dats_const) for (a, w) in wc_pairs]
        if not pairs:
            logger.info("Tidak ada pasangan aktif ditemukan. Selesai.")
            logger.info(f"Log tersimpan di: {RUN_LOG_FILE}")
            return

    if args.show_pairs:
        subhr()
        logger.info("DAFTAR PASANGAN:")
        for a, w, d in pairs:
            logger.info(f"  - {a}:{w} @ {d}")
        subhr()
        logger.info("")

    # DB prep
    db = ensure_db_and_table()
    cur = db.cursor()

    # Set PERNR yang akan dicek ke Z_RFC_DISPLAY_NIK_CONF
    role_pernrs: Set[str] = set()

    # Purge global blacklist bila diminta
    if args.purge_blacklist and blacklist:
        total_purged = 0
        CHUNK = 100
        ids = list(blacklist)
        logger.info("[PURGE] Menghapus baris PERNR blacklist dari tabel ...")
        for i in range(0, len(ids), CHUNK):
            chunk = ids[i: i + CHUNK]
            q = (
                f"DELETE FROM `{TABLE}` WHERE `pernr` IN ("
                + ",".join(["%s"] * len(chunk))
                + ")"
            )
            try:
                cur.execute(q, chunk)
                total_purged += cur.rowcount
            except mysql.connector.Error as e:
                db.rollback()
                logger.info(f"[ERROR] Purge blacklist: {e}")
            else:
                db.commit()
        logger.info(
            f"[PURGE] Terhapus {total_purged} baris (PERNR blacklist) dari seluruh tabel."
        )

    total_rows = 0
    total_deleted = 0
    summary: List[Tuple[str, str, int, int, int, float]] = []

    for (arbpl, werks, dats) in pairs:
        start = time.perf_counter()

        logger.info(f"PAIR  : ARBPL={arbpl} | WERKS={werks} | DATE={dats}")

        # 1) Ambil dari RFC
        if args.verbose_steps:
            logger.info("  MENGAMBIL DATA ...")
        try:
            rows = rows_from_rfc(rfc, arbpl, werks, dats)
        except (ABAPApplicationError, ABAPRuntimeError, CommunicationError) as e:
            logger.info(f"  [ERROR] RFC: {e}")
            summary.append((arbpl, werks, 0, 0, 0, time.perf_counter() - start))
            continue

        norm = [normalize_row(r, arbpl, werks) for r in rows]
        if args.verbose_steps:
            logger.info(f"    -> {len(norm)} baris dari RFC.")

        # Filter blacklist
        if blacklist:
            before = len(norm)
            norm = [r for r in norm if r.get("pernr") not in blacklist]
            skipped = before - len(norm)
            if skipped and args.verbose_steps:
                logger.info(f"  BLACKLIST    : {skipped} baris di-skip (PERNR)")

        # Kumpulkan PERNR untuk cek role
        for r in norm:
            p = (r.get("pernr") or "").strip()
            if p:
                role_pernrs.add(p)

        # Hitung bakal dihapus
        cur.execute(COUNT_SQL, (arbpl, werks))
        would_delete = int(cur.fetchone()[0])

        # 2) Hapus lama
        if args.no_delete:
            deleted = 0
            if args.verbose_steps:
                logger.info("  MENGHAPUS DATA LAMA : [LEWATI] (--no-delete)")
        elif args.dry_run:
            deleted = would_delete
            if args.verbose_steps:
                logger.info(
                    f"  MENGHAPUS DATA LAMA : [DRY-RUN] {would_delete} baris (simulasi)"
                )
        else:
            try:
                cur.execute(DELETE_SQL, (arbpl, werks))
                deleted = cur.rowcount
                db.commit()
                if args.verbose_steps:
                    logger.info(f"  MENGHAPUS DATA LAMA : {deleted} baris.")
            except mysql.connector.Error as e:
                db.rollback()
                logger.info(f"  [ERROR] DB DELETE: {e}")
                deleted = 0

        # 3) Insert baru
        inserted_for_pair = 0
        if args.dry_run:
            inserted_for_pair = len(norm)
            if args.verbose_steps:
                logger.info(
                    f"  MEMASUKKAN DATA BARU : [DRY-RUN] {len(norm)} baris (simulasi)"
                )
        elif norm:
            try:
                for i in range(0, len(norm), args.batch):
                    chunk = norm[i: i + args.batch]
                    cur.executemany(UPSERT_SQL, chunk)
                    inserted_for_pair += len(chunk)
                db.commit()
                if args.verbose_steps:
                    logger.info(
                        f"  MEMASUKKAN DATA BARU : {inserted_for_pair} baris."
                    )
            except mysql.connector.Error as e:
                db.rollback()
                logger.info(f"  [ERROR] DB INSERT: {e}")
                inserted_for_pair = 0
        else:
            if args.verbose_steps:
                logger.info("  MEMASUKKAN DATA BARU : 0 baris (tidak ada data).")

        elapsed = time.perf_counter() - start
        summary.append(
            (arbpl, werks, deleted, len(norm), inserted_for_pair, elapsed)
        )
        total_rows += inserted_for_pair
        total_deleted += deleted

    # Tutup cursor utama
    cur.close()

    # Cek role NIK via RFC kedua (kecuali dry-run)
    if not args.dry_run and role_pernrs:
        try:
            check_and_update_roles(rfc, db, role_pernrs, role_value="INDUK")
        except Exception as e:
            logger.info(f"[WARN] Gagal update role dari {RFC_ROLE_NAME}: {e}")

    # Setelah role selesai, update deskripsi WC
    if not args.dry_run:
        try:
            update_wc_descriptions(rfc, db, pairs)
        except Exception as e:
            logger.info(f"[WARN] Gagal update deskripsi WC dari {RFC_DESC_NAME}: {e}")

    # Setelah desc, update DEVISI dari Excel
    if not args.dry_run:
        try:
            devisi_mapping = load_devisi_mapping_from_excel(DEVISI_FILE)
            update_devisi_from_excel(db, pairs, devisi_mapping)
        except Exception as e:
            logger.info(f"[WARN] Gagal update devisi dari Excel: {e}")

    # Backup bulanan (hanya tanggal 6)
    if not args.dry_run:
        try:
            do_monthly_backup_if_needed(db)
        except Exception as e:
            logger.info(f"[WARN] Gagal melakukan backup bulanan: {e}")

    db.close()

    try:
        rfc.close()
    except Exception:
        pass

    # ---- RINGKASAN
    logger.info("")
    hr("=")
    logger.info("RINGKASAN")
    logger.info("ARBPL  | WERKS | Dihapus | Dari RFC | Insert/Update | Detik")
    subhr()
    for a, w, dl, rf, ins, sec in summary:
        logger.info(f"{a:<6} | {w:<5} | {dl:>7} | {rf:>8} | {ins:>13} | {sec:>5.2f}")
    subhr()
    logger.info(
        f"TOTAL  : di-insert/di-update={total_rows}, dihapus={total_deleted}."
    )
    logger.info(f"TABEL  : {DB_NAME}.{TABLE} (utf8mb4_unicode_ci).")
    hr("=")

    job_end_wall = datetime.datetime.now().astimezone()
    total_elapsed = time.perf_counter() - job_t0
    logger.info(f"MULAI   : {job_start_wall.strftime('%Y-%m-%d %H:%M:%S %Z%z')}")
    logger.info(f"SELESAI : {job_end_wall.strftime('%Y-%m-%d %H:%M:%S %Z%z')}")
    logger.info(f"DURASI  : {total_elapsed:.2f} detik")
    hr("=")

    logger.info(f"Main log (rotating) : {MAIN_LOG_FILE}")
    logger.info(f"Run log (unik)      : {RUN_LOG_FILE}")


if __name__ == "__main__":
    main()
