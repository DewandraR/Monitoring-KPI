#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# app_yppr058_refresh.py
# Flask API to refresh yppr058_data (Detail) AND wc_person_data (Master)
# Host: 0.0.0.0 Port: 5010
"""
export WC_BLACKLIST_FILE="/path/ke/blacklist_pernr.csv"
"""

import os, re, json, datetime, calendar, csv
import time
from decimal import Decimal
from typing import Any, Dict, List, Tuple, Optional
from pathlib import Path
from threading import Lock

from dotenv import load_dotenv
from flask import Flask, request, jsonify
from flask_cors import CORS

from pyrfc import (
    Connection,
    CommunicationError,
    LogonError,
    ABAPApplicationError,
    ABAPRuntimeError,
)
import mysql.connector
import logging
from logging.handlers import RotatingFileHandler

# Excel reader for DEVISI
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# ---------------- Env & Logging ----------------
load_dotenv()

PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_LOG_DIR = os.environ.get(
    "YPPR058_LOG_DIR",
    str(PROJECT_ROOT / "storage" / "logs" / "python_wc_person_mysql"),
)
os.makedirs(DEFAULT_LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(DEFAULT_LOG_DIR, "api_refresh.log")

# Lokasi file Excel DEVISI (Sesuaikan path ini dengan struktur folder Anda)
# Default: folder project root / DEVISI.xlsx
DEVISI_FILE = Path(os.environ.get("WC_DEVISI_FILE", str(PROJECT_ROOT / "DEVISI.xlsx")))

logger = logging.getLogger("yppr058_api")
logger.setLevel(logging.INFO)
fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch.setFormatter(fmt)

fh = RotatingFileHandler(LOG_FILE, maxBytes=2_000_000, backupCount=5, encoding="utf-8")
fh.setLevel(logging.INFO)
fh.setFormatter(fmt)

if not logger.handlers:
    logger.addHandler(ch)
    logger.addHandler(fh)

# ---------------- Filter untuk menghilangkan akses log /api/yppr058/progress -----

class IgnoreProgressAccessLog(logging.Filter):
    """
    Filter agar akses GET /api/yppr058/progress tidak muncul di log.
    Berlaku untuk logger 'werkzeug' dan root logger.
    """
    def filter(self, record: logging.LogRecord) -> bool:
        msg = record.getMessage()
        # Baris akses log biasanya berisi "GET /api/yppr058/progress?job_id=..."
        if "/api/yppr058/progress" in msg:
            return False   # jangan ditulis ke log
        return True        # request lain tetap ditulis

# Satu instance filter yang dipakai di beberapa logger/handler
ignore_progress_filter = IgnoreProgressAccessLog()

# Logger bawaan Flask / dev server (werkzeug)
werkzeug_logger = logging.getLogger("werkzeug")
werkzeug_logger.addFilter(ignore_progress_filter)
for handler in list(werkzeug_logger.handlers):
    handler.addFilter(ignore_progress_filter)

# ROOT logger – sering jadi tempat akses log mendarat
root_logger = logging.getLogger()
root_logger.addFilter(ignore_progress_filter)
for handler in list(root_logger.handlers):
    handler.addFilter(ignore_progress_filter)

# ---------------- PROGRESS (untuk /api/yppr058/refresh) ----------------
# Menyimpan progres job refresh detail per job_id (in-memory)
PROGRESS_MAP: Dict[str, Dict[str, Any]] = {}
PROGRESS_LOCK = Lock()

# ---------------- SAP ----------------
DEFAULT_SAP = {
    "ashost": os.environ.get("SAP_ASHOST", "192.168.254.154"),
    "sysnr": os.environ.get("SAP_SYSNR", "01"),
    "client": os.environ.get("SAP_CLIENT", "300"),
    "lang": os.environ.get("SAP_LANG", "EN"),
}
SAP_USERNAME = os.environ.get("SAP_USER", "auto_email")
SAP_PASSWORD = os.environ.get("SAP_PASS", "11223344")

# RFC NAMES
RFC_NAME = "Z_FM_YPPR058DX"               # Detail Transaksi
RFC_MAIN_WC = "CR_PERSONS_OF_WORKCENTER"  # Master WC
RFC_ROLE_WC = "Z_RFC_DISPLAY_NIK_CONF"    # Role Induk
RFC_DESC_WC = "Z_FM_GET_WC_DESC"          # Deskripsi WC
RFC_LOG_MUTA = "Z_FM_YPP_LOG_MUTA"       # Log Mutasi (untuk YPPR058)

# ---------------- MySQL ----------------
DB_HOST = os.environ.get("DB_HOST", "127.0.0.1")
DB_PORT = int(os.environ.get("DB_PORT", "3306"))
DB_USER = os.environ.get("DB_USER", "root")
DB_PASS = os.environ.get("DB_PASS", "root")
DB_NAME = os.environ.get("DB_NAME", "wc_person")

OUT_TABLE = os.environ.get("DB_TABLE_OUT", "yppr058_data")
WC_TABLE = os.environ.get("WC_TABLE", "wc_person_data")

# ---------------- SQL QUERY: YPPR058 (Detail) ----------------
ALLOW_EMPTY_DELETE = str(os.environ.get("ALLOW_EMPTY_DELETE", "false")).lower() in ("1", "true", "yes")

CHECK_CONFIRM_SQL = f"""
SELECT COUNT(*) FROM `{OUT_TABLE}`
WHERE pernr = %s AND begda = %s AND arbpl2 IS NOT NULL AND arbpl2 <> ''
"""

FIND_INDUK_SQL = f"""
SELECT DISTINCT pernr FROM `{WC_TABLE}`
WHERE arbpl = %s AND werks = %s AND begda <= %s AND endda >= %s
  AND (role = 'INDUK' OR role = 'Induk' OR role = 'induk')
"""

UPSERT_SQL_YPPR = f"""
INSERT INTO `{OUT_TABLE}`
(`pernr`,`begda`,`total_jam`,`mint2`,`mintu`,`mintu2`,`mintu3`,
 `cname`,`gji`,`gji2`,`varnt`,`varnt1`,
 `arbpl`,`arbpl2`,`shift`,`werks`,`desc`,`role`,`devisi`,`source_rfc`)
VALUES (%(pernr)s,%(begda)s,%(total_jam)s,%(mint2)s,%(mintu)s,%(mintu2)s,%(mintu3)s,
        %(cname)s,%(gji)s,%(gji2)s,%(varnt)s,%(varnt1)s,
        %(arbpl)s,%(arbpl2)s,%(shift)s,%(werks)s,%(desc)s,%(role)s,%(devisi)s,'{RFC_NAME}')
ON DUPLICATE KEY UPDATE
  `total_jam`=VALUES(`total_jam`),
  `mint2`=VALUES(`mint2`),
  `mintu`=VALUES(`mintu`),
  `mintu2`=VALUES(`mintu2`),
  `mintu3`=VALUES(`mintu3`),
  `cname`=VALUES(`cname`),
  `gji`=VALUES(`gji`),
  `gji2`=VALUES(`gji2`),
  `varnt`=VALUES(`varnt`),
  `varnt1`=VALUES(`varnt1`),
  `arbpl2`=VALUES(`arbpl2`),
  `shift`=VALUES(`shift`),
  `desc`=VALUES(`desc`),
  `role`=VALUES(`role`),
  `devisi`=VALUES(`devisi`),
  `inserted_at`=CURRENT_TIMESTAMP
"""

# ---------------- SQL QUERY: WC PERSON (Master) ----------------
DELETE_WC_SQL = f"DELETE FROM `{WC_TABLE}` WHERE `arbpl`=%s AND `werks`=%s"
COUNT_WC_SQL  = f"SELECT COUNT(*) FROM `{WC_TABLE}` WHERE `arbpl`=%s AND `werks`=%s"

DELETE_YPPR_WC_SQL = f"""
DELETE FROM `{OUT_TABLE}`
WHERE `arbpl` = %s
  AND `werks` = %s
"""

# NEW: ambil daftar PERNR lama per WC/Plant (untuk deteksi NIK baru)
GET_OLD_PERNRS_SQL = f"""
SELECT DISTINCT pernr FROM `{WC_TABLE}`
WHERE `arbpl`=%s AND `werks`=%s
"""

UPSERT_WC_SQL = f"""
INSERT INTO `{WC_TABLE}`
(`otype`,`objid`,`pernr`,`begda`,`endda`,`arbid`,`short`,`stext`,`arbpl`,`werks`,`source_rfc`)
VALUES (%(otype)s,%(objid)s,%(pernr)s,%(begda)s,%(endda)s,%(arbid)s,%(short)s,%(stext)s,%(arbpl)s,%(werks)s,'{RFC_MAIN_WC}')
ON DUPLICATE KEY UPDATE
  `endda`=VALUES(`endda`),
  `short`=VALUES(`short`),
  `stext`=VALUES(`stext`),
  `arbpl`=VALUES(`arbpl`),
  `werks`=VALUES(`werks`),
  `inserted_at`=CURRENT_TIMESTAMP
"""

UPDATE_ROLE_WC_SQL = f"UPDATE `{WC_TABLE}` SET `role`=%s WHERE `pernr`=%s"
UPDATE_DESC_WC_SQL = f"UPDATE `{WC_TABLE}` SET `desc`=%s WHERE `arbpl`=%s AND `werks`=%s"
UPDATE_EXTRA_INFO_SQL = f"UPDATE `{WC_TABLE}` SET `devisi`=%s, `kode_laravel`=%s WHERE `arbpl`=%s AND `werks`=%s"

# ---------------- Helpers umum ----------------
def to_dats(s: str) -> str:
    s = (s or "").strip()
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", s)
    if m:
        return f"{m.group(3)}{m.group(2)}{m.group(1)}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return f"{m.group(1)}{m.group(2)}{m.group(3)}"
    m = re.match(r"^\d{8}$", s)
    if m:
        return s
    raise ValueError(f"Invalid date format: {s}")


def month_range(dats: str) -> Tuple[str, str]:
    y = int(dats[0:4])
    m = int(dats[4:6])
    start = f"{y:04d}{m:02d}01"
    last = calendar.monthrange(y, m)[1]
    end = f"{y:04d}{m:02d}{last:02d}"
    return start, end


def get_mysql():
    return mysql.connector.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS,
        database=DB_NAME, autocommit=False,
    )


def acquire_pair_lock(cur, arbpl: str, werks: str, timeout: int = 120) -> bool:
    cur.execute("SELECT GET_LOCK(CONCAT('yppr058:', %s, ':', %s), %s)", (arbpl, werks, timeout))
    row = cur.fetchone()
    return bool(row and row[0] == 1)


def release_pair_lock(cur, arbpl: str, werks: str):
    try:
        cur.execute("SELECT RELEASE_LOCK(CONCAT('yppr058:', %s, ':', %s))", (arbpl, werks))
        cur.fetchone()
    except Exception:
        pass


def norm_val(x):
    if x is None:
        return None
    if isinstance(x, Decimal):
        return x
    try:
        return Decimal(str(x)) if isinstance(x, (int, float)) else x
    except Exception:
        return x

# --- Helpers untuk YPPR058 (Detail) ---
def get_wc_desc(conn, arbpl: str, werks: str) -> Optional[str]:
    sql = f"SELECT `desc` FROM `{WC_TABLE}` WHERE arbpl=%s AND werks=%s AND `desc` IS NOT NULL AND `desc`<>'' LIMIT 1"
    cur = conn.cursor()
    try:
        cur.execute(sql, (arbpl, werks))
        row = cur.fetchone()
        return row[0] if row else None
    except Exception:
        return None
    finally:
        cur.close()


def get_wc_meta_for_pair(conn, arbpl: str, werks: str, dats: str) -> Tuple[Optional[str], Optional[str], Dict[str, str]]:
    """
    Ambil metadata WC dari wc_person_data untuk satu pasangan (arbpl, werks) & tanggal:
    - desc_val  : deskripsi WC
    - devisi_val: nama devisi
    - role_map  : mapping pernr -> role (INDUK/dll)
    """
    desc_val: Optional[str] = None
    devisi_val: Optional[str] = None
    role_map: Dict[str, str] = {}

    cur = conn.cursor()
    try:
        sql = f"""
            SELECT pernr, role, devisi, `desc`
            FROM `{WC_TABLE}`
            WHERE arbpl=%s AND werks=%s
              AND begda <= %s AND endda >= %s
        """
        cur.execute(sql, (arbpl, werks, dats, dats))
        for pernr, role, devisi, desc_ in cur.fetchall():
            p = str(pernr or "").strip()
            if p.isdigit():
                p = p.zfill(8)

            if desc_val is None and desc_:
                desc_val = str(desc_)

            if devisi_val is None and devisi:
                devisi_val = str(devisi)

            if role:
                role_map.setdefault(p, str(role).strip())
    except Exception as e:
        logger.error(f"[YPPR058] get_wc_meta_for_pair error for {arbpl}/{werks}@{dats}: {e}")
    finally:
        cur.close()

    return desc_val, devisi_val, role_map


def normalize_tdata(
    row: Dict[str, Any],
    arbpl: str,
    werks: str,
    desc_value: Optional[str],
    role_value: Optional[str] = None,
    devisi_value: Optional[str] = None,
) -> Dict[str, Any]:
    S = lambda v: "" if v is None else str(v)
    I = lambda v: None if (S(v) == "") else int(S(v))
    D = lambda v: None if v in (None, "") else norm_val(v)

    pernr_raw = S(row.get("PERNR"))
    pernr_norm = pernr_raw.zfill(8) if pernr_raw.isdigit() else pernr_raw

    return {
        "pernr": pernr_norm,
        "begda": S(row.get("BEGDA")),
        "total_jam": D(row.get("TOTAL_JAM")),
        "mint2": I(row.get("MINT2")),
        "mintu": I(row.get("MINTU")),
        "mintu2": I(row.get("MINTU2")),
        "mintu3": I(row.get("MINTU3")),
        "cname": S(row.get("CNAME")),
        "gji": D(row.get("GJI")),
        "gji2": D(row.get("GJI2")),
        "varnt": D(row.get("VARNT")),
        "varnt1": D(row.get("VARNT1")),
        "arbpl": S(row.get("ARBPL")) or arbpl,
        "arbpl2": S(row.get("ARBPL2")),
        "shift": I(row.get("SHIFT")),
        "werks": werks,
        "desc": desc_value,
        "role": (S(role_value) or None),
        "devisi": (S(devisi_value) or None),
    }

# --- Helpers untuk WC PERSON (Master) ---
def normalize_wc_row(r: Dict, arbpl: str, werks: str) -> Dict:
    S = lambda x: "" if x is None else str(x).strip()
    return {
        "otype": S(r.get("OTYPE")),
        "objid": S(r.get("OBJID")),
        "pernr": S(r.get("PERNR")).zfill(8) if S(r.get("PERNR")).isdigit() else S(r.get("PERNR")),
        "begda": S(r.get("BEGDA")),
        "endda": S(r.get("ENDDA")),
        "arbid": S(r.get("ARBID")),
        "short": S(r.get("SHORT")),
        "stext": S(r.get("STEXT")),
        "arbpl": S(r.get("ARBPL")) or arbpl,
        "werks": S(r.get("WERKS")) or werks,
    }

# --- EXCEL DEVISI HELPER (New for API) ---
def load_devisi_mapping_from_excel(path: Path) -> Dict[Tuple[str, str], Dict[str, str]]:
    """
    Baca Excel:

    - Kolom B (idx 1): PLANT
    - Kolom C (idx 2): DEVISI
    - Kolom E (idx 4): KODE (ARBPL)
    - Kolom F (idx 5): KODE LARAVEL (bisa merged ke beberapa baris)
    """
    mapping: Dict[Tuple[str, str], Dict[str, str]] = {}

    if load_workbook is None:
        logger.warning("[EXCEL] openpyxl not installed. Skipping excel update.")
        return mapping

    p = Path(path)
    if not p.is_file():
        logger.warning(f"[EXCEL] File not found: {p}")
        return mapping

    try:
        wb = load_workbook(filename=str(p), data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"[EXCEL] Error opening Excel: {e}")
        return mapping

    # --- BACA MERGED CELL UNTUK KOLOM F (KODE LARAVEL) ---
    # mapping: row_index -> nilai kode_laravel hasil merge
    merged_kolar_by_row: Dict[int, str] = {}
    for cell_range in ws.merged_cells.ranges:
        # Hanya kalau range menyentuh kolom F (kolom ke-6)
        if cell_range.min_col <= 6 <= cell_range.max_col:
            top_cell = ws.cell(row=cell_range.min_row, column=6)
            if top_cell.value not in (None, ""):
                val = str(top_cell.value).strip()
                for r in range(cell_range.min_row, cell_range.max_row + 1):
                    merged_kolar_by_row[r] = val

    current_devisi = ""
    row_mapped = 0

    # Asumsi: Row 1 header, data mulai baris 2
    for row in ws.iter_rows(min_row=2):
        plant_cell = row[1] if len(row) > 1 else None  # B
        devisi_cell = row[2] if len(row) > 2 else None  # C
        kode_cell   = row[4] if len(row) > 4 else None  # E
        kolar_cell  = row[5] if len(row) > 5 else None  # F

        plant_val = ""
        kode_val  = ""

        # PLANT (B)
        if plant_cell and plant_cell.value is not None:
            plant_val = str(plant_cell.value).strip()

        # DEVISI (C) – mekanisme "last non-empty"
        if devisi_cell and devisi_cell.value not in (None, ""):
            current_devisi = str(devisi_cell.value).strip()

        # KODE / ARBPL (E)
        if kode_cell and kode_cell.value is not None:
            kode_val = str(kode_cell.value).strip()

        # KODE LARAVEL (F)
        kode_laravel_val = ""
        if kolar_cell:
            if kolar_cell.value not in (None, ""):
                # Ada nilai langsung di sel F
                kode_laravel_val = str(kolar_cell.value).strip()
            else:
                # Cek apakah baris ini bagian dari merged cell kolom F
                merged_val = merged_kolar_by_row.get(kolar_cell.row)
                if merged_val is not None:
                    kode_laravel_val = merged_val
                # Kalau tidak ada merged_val -> biarkan kosong ("") => nanti jadi NULL di DB

        # Kalau plant/kode kosong → skip
        if not plant_val or not kode_val:
            continue

        key = (kode_val.upper(), plant_val)
        if key not in mapping:
            mapping[key] = {
                "devisi": current_devisi,
                "kode_laravel": kode_laravel_val,
            }
            row_mapped += 1

    logger.info(f"[EXCEL] Total mapping loaded: {row_mapped}")
    return mapping

# ---------------- Blacklist PERNR (NIK) ----------------
HARDCODED_BLACKLIST = {
    "10000011",
    "10000015",
    "10000040",
    "10000062",
    "10000063",
    "10000083",
    "10000110",
    "10000126",
    "10000144",
    "10000161",
    "10000189",
    "10000364",
    "10000395",
    "10000414",
    "10000417",
    "10000427",
    "10000431",
    "10000440",
    "10000458",
    "10000482",
    "10000502",
    "10000524",
    "10000541",
    "10000548",
    "10000555",
    "10000564",
    "10000570",
    "10000577",
    "10000591",
    "10000615",
    "10000622",
    "10000642",
    "10000659",
    "10000725",
    "10000778",
    "10000874",
    "10001561",
    "10001983",
    "10002308",
    "10002690",
    "10002787",
    "10003007",
    "10003008",
    "10003009",
    "10004934",
    "10004994",
    "10005063",
    "10003590",
    "10003599",
    "10003600",
    "10004874",
    "10000897",
    "10000898",
    "10001002",
    "10006163",
    "10006337",
    "10007161",
    "10002446",
    "10000467",
    "10004411",
    "10000644",
    "10000026",
    "10000093",
    "10000109",
    "10000112",
    "10000141",
    "10000266",
    "10000319",
    "10002804",
    "10000420",
    "10005689",
    "10008126",
    "10008135",
    "10008134",
}

def _pad8(s: str) -> str:
    s = (s or "").strip()
    return s.zfill(8) if s.isdigit() else s

def load_blacklist(extra_file: Optional[str] = None) -> set:
    """
    Gabung:
    - HARDCODED_BLACKLIST
    - ENV WC_BLACKLIST_PERNR = "10001234,10005678"
    - file (opsional, ENV: WC_BLACKLIST_FILE), boleh TXT/CSV kolom PERNR
    """
    bl = {_pad8(x) for x in HARDCODED_BLACKLIST}

    # Dari ENV (comma separated)
    env_val = os.environ.get("WC_BLACKLIST_PERNR", "")
    for tok in [t.strip() for t in env_val.split(",") if t.strip()]:
        bl.add(_pad8(tok))

    # Dari file (opsional)
    if extra_file:
        try:
            with open(extra_file, "r", encoding="utf-8") as f:
                head = f.read(2048)
                f.seek(0)

                # Mode CSV dengan kolom PERNR
                if "," in head and "pernr" in head.lower():
                    rdr = csv.DictReader(f)
                    for r in rdr:
                        for k in r.keys():
                            if k and k.lower() == "pernr":
                                bl.add(_pad8(str(r[k])))
                else:
                    # Mode TXT: pisah spasi/comma/semicolon
                    for line in f:
                        for tok in re.split(r"[\s,;]+", line.strip()):
                            if tok:
                                bl.add(_pad8(tok))
        except Exception as e:
            logger.warning(f"[BLACKLIST] gagal baca file blacklist: {e}")

    logger.info(f"[BLACKLIST] aktif: {len(bl)} PERNR")
    return bl

def is_blacklisted(pernr: str) -> bool:
    return _pad8(pernr) in BLACKLIST_PERNRS

# Inisialisasi sekali saat app start
BLACKLIST_PERNRS = load_blacklist(os.environ.get("WC_BLACKLIST_FILE", ""))

# ---------------- PROGRESS helpers ----------------
def init_progress(job_id: str, total_items: int):
    """
    Inisialisasi progres untuk job refresh detail.
    total_items = jumlah items (kombinasi pernr+tanggal) yang akan diproses.
    """
    with PROGRESS_LOCK:
        PROGRESS_MAP[job_id] = {
            "job_id": job_id,
            "total_items": int(total_items),
            "done_items": 0,
            "status": "running",
            "started_at": datetime.datetime.now().isoformat(),
            "finished_at": None,
            "last_item": None,
        }

def update_progress(
    job_id: str,
    done_items: int,
    *,
    pernr: Optional[str] = None,
    begda: Optional[str] = None,
    arbpl: Optional[str] = None,
    werks: Optional[str] = None,
):
    """
    Update progres job_id: berapa item yang sudah selesai, plus info item terakhir.
    done_items = index item terakhir yang sudah diproses (1-based).
    """
    with PROGRESS_LOCK:
        info = PROGRESS_MAP.get(job_id)
        if not info:
            return
        info["done_items"] = int(done_items)
        if any([pernr, begda, arbpl, werks]):
            info["last_item"] = {
                "pernr": pernr,
                "begda": begda,
                "arbpl": arbpl,
                "werks": werks,
                "time": datetime.datetime.now().isoformat(),
            }

def finish_progress(job_id: str, ok: bool = True, error: Optional[str] = None):
    """
    Tandai job selesai (ok / error).
    """
    with PROGRESS_LOCK:
        info = PROGRESS_MAP.get(job_id)
        if not info:
            return
        info["status"] = "done" if ok else "error"
        info["finished_at"] = datetime.datetime.now().isoformat()
        if error:
            info["error"] = str(error)

# ---------------- Flask App ----------------
app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

@app.get("/health")
def health():
    return {"ok": True, "service": "yppr058_refresh", "time": datetime.datetime.now().isoformat()}

# ---------------- Endpoint progress untuk frontend ----------------
@app.get("/api/yppr058/progress")
def api_refresh_progress():
    """
    Query:
      GET /api/yppr058/progress?job_id=xxx

    Response:
      {
        "ok": true,
        "progress": {
          "job_id": "...",
          "total_items": 23,
          "done_items": 8,
          "status": "running|done|error",
          "percent_items": 34.8,
          "last_item": {...}
        }
      }
    """
    job_id = (request.args.get("job_id") or "").strip()
    if not job_id:
        return jsonify({"ok": False, "error": "job_id required"}), 400

    with PROGRESS_LOCK:
        info = PROGRESS_MAP.get(job_id)

    if not info:
        return jsonify({"ok": False, "error": "job not found"}), 404

    total = max(1, int(info.get("total_items") or 1))
    done = int(info.get("done_items") or 0)
    percent = round(done * 100.0 / total, 1)

    payload = dict(info)
    payload["percent_items"] = percent

    return jsonify({"ok": True, "progress": payload})

# ============================================================================
# ENDPOINT 1: REFRESH DETAIL (YPPR058)
# ============================================================================
def delete_old_rows(cur, arbpl: str, werks: str, begda: str, endda: str, pernrs: List[str]) -> int:
    if not pernrs:
        return 0
    placeholders = ",".join(["%s"] * len(pernrs))
    sql = f"DELETE FROM `{OUT_TABLE}` WHERE `arbpl`=%s AND `werks`=%s AND `begda` BETWEEN %s AND %s AND `pernr` IN ({placeholders})"
    params = [arbpl, werks, begda, endda] + list(pernrs)
    cur.execute(sql, params)
    return cur.rowcount

def call_rfc_yppr(conn: Connection, arbpl: str, werks: str, begda: str, endda: str, pernrs: List[str]):
    return conn.call(
        RFC_NAME, P_BEGDA=begda, P_ENDDA=endda, P_WERKS=werks, P_ARBPL=arbpl,
        T_ARBPL=[{"ARBPL": arbpl}], T_PERNR=[{"PERNR": p} for p in pernrs],
    )


def call_rfc_log_muta(conn: Connection, pernr: str, begda: str, endda: str) -> Tuple[Optional[int], Optional[str]]:
    """
    Panggil Z_FM_YPP_LOG_MUTA untuk 1 PERNR & range tanggal.
    Return: (EV_SUBRC, EV_MSG). Jika error -> (None, pesan_error).

    Aturan pakai di /api/yppr058/refresh:
    - EV_SUBRC == 0  -> PERNR diproses SOLO (tanpa induk).
    - EV_SUBRC != 0  -> PERNR ikut grup (boleh ajak induk).
    """
    try:
        resp = conn.call(
            RFC_LOG_MUTA,
            IV_BEGDA=begda,
            IV_ENDDA=endda,
            IV_PERNR=pernr,
        )
        subrc_raw = resp.get("EV_SUBRC")
        msg = (resp.get("EV_MSG") or "").strip()

        try:
            subrc = int(subrc_raw)
        except Exception:
            subrc = None

        logger.info(
            f"[LOG_MUTA] PERNR={pernr} {begda}..{endda} "
            f"EV_SUBRC={subrc} MSG='{msg}'"
        )
        return subrc, msg

    except (ABAPApplicationError, ABAPRuntimeError, CommunicationError) as e:
        logger.error(
            f"[LOG_MUTA] RFC error for {pernr}@{begda}..{endda}: {e}"
        )
        return None, str(e)

def resolve_pairs(cur, pernr: str, dats: str) -> Tuple[str, List[Tuple[str, str]]]:
    cur.execute(
        f"SELECT DISTINCT arbpl, werks FROM `{WC_TABLE}` "
        f"WHERE pernr=%s AND begda<=%s AND endda>=%s AND arbpl<>'' AND werks<>''",
        (pernr, dats, dats),
    )
    rows = [(a or "", w or "") for (a, w) in cur.fetchall()]
    if rows:
        return "wc_person", rows
    m_start, m_end = month_range(dats)
    cur.execute(
        f"SELECT arbpl, werks, COUNT(*) c, MAX(begda) last_d "
        f"FROM `{OUT_TABLE}` "
        f"WHERE pernr=%s AND begda BETWEEN %s AND %s AND arbpl<>'' AND werks<>'' "
        f"GROUP BY arbpl, werks ORDER BY c DESC, last_d DESC LIMIT 3",
        (pernr, m_start, m_end),
    )
    rows = [(a or "", w or "") for (a, w, *_rest) in cur.fetchall()]
    if rows:
        return "yppr058_recent", rows
    return "none", []

def pernr_has_confirm(cur, pernr: str, dats: str) -> bool:
    cur.execute(CHECK_CONFIRM_SQL, (pernr, dats))
    row = cur.fetchone()
    return bool(row and row[0] > 0)

def find_induk_for_pair(cur, arbpl: str, werks: str, dats: str) -> List[str]:
    cur.execute(FIND_INDUK_SQL, (arbpl, werks, dats, dats))
    return [str(p).strip().zfill(8) for (p,) in cur.fetchall() if p]

@app.post("/api/yppr058/refresh")
def api_refresh_detail():
    """
    Body:
    {
      "items": [
        {"pernr":"10005817","werks":"","arbpl":"","begda":"20251108","endda":"20251108"},
        ...
      ],
      "job_id": "opsional-string-unik"
    }

    items = kombinasi (pernr, begda, endda, arbpl, werks) yang akan diproses.
    Satu item = satu kombinasi NIK-tanggal.

    MODIFIKASI:
    - Sebelum tarik YPPR058, setiap PERNR di 'base_pernrs' dicek ke Z_FM_YPP_LOG_MUTA:
        EV_SUBRC == 0  -> dimasukkan ke solo_pernrs (ditarik sendiri, tanpa induk).
        EV_SUBRC != 0  -> dimasukkan ke group_pernrs (boleh mengajak induk).
    - Grup yang sama (WC/Plant sama) tetap di 1 paket, hanya yang EV_SUBRC=0 dipisah.
    """

    # --- Baca JSON ---
    try:
        payload = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"ok": False, "error": "Invalid JSON"}), 400

    items = payload.get("items") or []
    if not isinstance(items, list) or not items:
        return jsonify({"ok": False, "error": "items[] required"}), 400

    # job_id dari frontend (atau auto-generate)
    job_id = str(payload.get("job_id") or "").strip()
    if not job_id:
        job_id = f"job-{int(time.time() * 1000)}"

    # inisialisasi progres
    init_progress(job_id, len(items))

    # --- SAP connect sekali per request ---
    sap_params = {**DEFAULT_SAP, "user": SAP_USERNAME, "passwd": SAP_PASSWORD}
    try:
        rfc = Connection(**sap_params)
    except (CommunicationError, LogonError) as e:
        logger.error(f"SAP connect failed: {e}")
        finish_progress(job_id, ok=False, error=f"SAP connect failed: {e}")
        return jsonify({"ok": False, "error": f"SAP connect failed: {e}", "job_id": job_id}), 500

    db = get_mysql()
    cur = db.cursor()

    results: List[Dict[str, Any]] = []

    try:
        for idx, it in enumerate(items, start=1):
            # ------------------------------
            # 1) Ambil daftar NIK dasar (multi)
            # ------------------------------
            base_pernrs: List[str] = []
            raw_pernrs = it.get("pernrs")

            if isinstance(raw_pernrs, list) and raw_pernrs:
                for p in raw_pernrs:
                    p_str = str(p or "").strip()
                    if not p_str:
                        continue
                    if p_str.isdigit():
                        p_str = p_str.zfill(8)

                    if is_blacklisted(p_str):
                        logger.info(f"[REFRESH] Skip {p_str} (multi) karena BLACKLIST.")
                        continue

                    if p_str not in base_pernrs:
                        base_pernrs.append(p_str)

            # 2) Fallback: kalau tidak ada 'pernrs', pakai 'pernr' biasa (perilaku lama)
            if not base_pernrs:
                pernr_single = str(it.get("pernr") or "").strip()
                if pernr_single:
                    if pernr_single.isdigit():
                        pernr_single = pernr_single.zfill(8)

                    if is_blacklisted(pernr_single):
                        logger.info(f"[REFRESH] Skip {pernr_single} karena BLACKLIST.")
                        results.append(
                            {
                                "ok": False,
                                "pernr": pernr_single,
                                "skipped": True,
                                "reason": "blacklisted",
                            }
                        )
                        update_progress(job_id, idx, pernr=pernr_single)
                        continue

                    base_pernrs.append(pernr_single)

            # Kalau masih kosong -> tidak ada NIK valid sama sekali
            if not base_pernrs:
                results.append(
                    {
                        "ok": False,
                        "pernr": None,
                        "error": "Missing pernr/pernrs",
                    }
                )
                update_progress(job_id, idx)
                continue

            # Untuk log/progress pakai NIK pertama
            pernr = base_pernrs[0]

            req_werks = (it.get("werks") or "").strip()
            req_arbpl = (it.get("arbpl") or "").strip()

            begda_for_progress: Optional[str] = None
            arbpl_for_progress: Optional[str] = None
            werks_for_progress: Optional[str] = None

            try:
                # --- Normalisasi tanggal ---
                try:
                    begda = to_dats(str(it.get("begda") or ""))
                    endda = to_dats(str(it.get("endda") or begda))
                    begda_for_progress = begda
                except Exception as e:
                    results.append({"ok": False, "pernr": pernr, "error": str(e)})
                    continue

                if not begda or not endda:
                    results.append(
                        {"ok": False, "pernr": pernr, "error": "Missing begda/endda"}
                    )
                    continue

                # ------------------------------------------------------
                # 2b) CEK LOG MUTASI (Z_FM_YPP_LOG_MUTA) PER NIK
                # ------------------------------------------------------
                solo_pernrs: List[str] = []
                group_pernrs: List[str] = []
                log_muta_info: Dict[str, Dict[str, Any]] = {}

                for p in base_pernrs:
                    subrc, msg = call_rfc_log_muta(rfc, p, begda, endda)
                    log_muta_info[p] = {"subrc": subrc, "msg": msg}

                    # Aturan:
                    #   EV_SUBRC == 0  -> diproses sendiri (solo, tanpa induk).
                    #   lainnya         -> tetap ikut grup (boleh ajak induk).
                    if subrc == 0:
                        solo_pernrs.append(p)
                    else:
                        group_pernrs.append(p)

                logger.info(
                    f"[LOG_MUTA] paket={base_pernrs} begda..endda={begda}..{endda} "
                    f"solo={solo_pernrs} group={group_pernrs}"
                )

                # ------------------------------------------------------
                # 3) Cek apakah di kelompok group_pernrs perlu tarik INDUK
                #     (kalau minimal ada 1 NIK yang belum punya WC konfirmasi)
                #    NIK yang solo TIDAK akan mengajak induk.
                # ------------------------------------------------------
                needs_induk_any = False
                for p in group_pernrs:
                    try:
                        if not pernr_has_confirm(cur, p, begda):
                            needs_induk_any = True
                            break
                    except Exception as e:
                        logger.error(f"[DB] gagal cek WC Confirmasi untuk {p}@{begda}: {e}")

                # --- Tentukan pasangan WC/Plant ---
                if req_arbpl and req_werks:
                    strategy, pairs = "request", [(req_arbpl, req_werks)]
                else:
                    # asumsi: kalau kamu pakai 'pernrs', semua NIK di paket ini memang 1 WC
                    strategy, pairs = resolve_pairs(cur, pernr, begda)

                if pairs:
                    arbpl_for_progress, werks_for_progress = pairs[0]

                logger.info(
                    f"[REQ] pernrs={base_pernrs} begda..endda={begda}..{endda} "
                    f"strategy={strategy} pairs={pairs} "
                    f"needs_induk_any(group)={needs_induk_any}"
                )

                if not pairs:
                    results.append(
                        {
                            "ok": False,
                            "pernr": pernr,
                            "strategy": strategy,
                            "pairs_used": [],
                            "deleted_total": 0,
                            "inserted_total": 0,
                            "error": "No WC/Plant pair found; skipped",
                        }
                    )
                    continue

                item_pairs_report: List[Dict[str, Any]] = []
                total_deleted = 0
                total_inserted = 0

                # ------------------------------------------------------
                # 4) Loop per pair WC/Plant
                #    - pertama proses group_pernrs (jika ada)
                #    - lalu proses masing-masing solo_pernrs (jika ada)
                # ------------------------------------------------------
                for (arbpl, werks) in pairs:
                    if not acquire_pair_lock(cur, arbpl, werks, timeout=120):
                        logger.warning(
                            f"[LOCK] Timeout lock yppr058:{arbpl}:{werks} "
                            f"untuk {base_pernrs}@{begda}"
                        )
                        item_pairs_report.append(
                            {
                                "arbpl": arbpl,
                                "werks": werks,
                                "ok": False,
                                "error": "Lock timeout",
                            }
                        )
                        continue

                    deleted_pair = 0
                    inserted_pair = 0

                    try:
                        # --- Ambil metadata WC dari wc_person_data (sekali per pair) ---
                        desc_val, devisi_val, role_map = get_wc_meta_for_pair(
                            db, arbpl, werks, begda
                        )

                        # ======================================================
                        # 4a) PROSES GRUP (group_pernrs) -> bisa ajak induk
                        # ======================================================
                        if group_pernrs:
                            pernrs_for_rfc: List[str] = list(group_pernrs)

                            # Jika butuh induk: tambahkan NIK induk
                            if needs_induk_any:
                                try:
                                    induks = find_induk_for_pair(cur, arbpl, werks, begda)
                                except Exception as e:
                                    logger.error(
                                        f"[DB] gagal cari induk untuk {group_pernrs}@"
                                        f"{arbpl}/{werks} {begda}: {e}"
                                    )
                                    induks = []

                                if induks:
                                    seen = set(pernrs_for_rfc)
                                    extra: List[str] = []
                                    for ip in induks:
                                        ip = str(ip or "").strip()
                                        if ip.isdigit():
                                            ip = ip.zfill(8)

                                        # Induk yang juga ada di solo_pernrs TIDAK diikutkan
                                        # (sesuai requirement: kalau EV_SUBRC=0 -> abaikan induk).
                                        if ip in solo_pernrs:
                                            logger.info(
                                                f"[ROLE] Skip induk {ip} karena termasuk solo_pernrs "
                                                f"untuk paket {base_pernrs}@{arbpl}/{werks}"
                                            )
                                            continue

                                        if ip in seen:
                                            continue
                                        if is_blacklisted(ip):
                                            logger.info(
                                                f"[REFRESH] Skip induk {ip} karena BLACKLIST "
                                                f"untuk {group_pernrs}@{arbpl}/{werks}"
                                            )
                                            continue

                                        seen.add(ip)
                                        extra.append(ip)

                                    if extra:
                                        pernrs_for_rfc.extend(extra)
                                        logger.info(
                                            f"[ROLE] group_pernrs={group_pernrs} perlu induk, "
                                            f"tambah induk={extra} untuk {arbpl}/{werks}"
                                        )

                            # Filter blacklist lagi (jaga-jaga)
                            if BLACKLIST_PERNRS:
                                before_bl = len(pernrs_for_rfc)
                                pernrs_for_rfc = [
                                    p for p in pernrs_for_rfc if not is_blacklisted(p)
                                ]
                                removed_bl = before_bl - len(pernrs_for_rfc)
                                if removed_bl:
                                    logger.info(
                                        f"[REFRESH] {removed_bl} PERNR (group+induk) di-skip karena BLACKLIST "
                                        f"untuk {group_pernrs}@{arbpl}/{werks}"
                                    )

                            if pernrs_for_rfc:
                                # --- CALL RFC (YPPR058DX) untuk grup ---
                                try:
                                    logger.info(
                                        f"[SAP CALL GROUP] {arbpl}/{werks} {begda}..{endda} "
                                        f"T_PERNR={pernrs_for_rfc}"
                                    )
                                    resp = call_rfc_yppr(
                                        rfc, arbpl, werks, begda, endda, pernrs_for_rfc
                                    )
                                except (ABAPApplicationError, ABAPRuntimeError, CommunicationError) as e:
                                    logger.error(f"[SAP] {e}")
                                    item_pairs_report.append(
                                        {
                                            "arbpl": arbpl,
                                            "werks": werks,
                                            "mode": "group",
                                            "ok": False,
                                            "error": f"SAP: {e}",
                                            "pernrs": pernrs_for_rfc,
                                        }
                                    )
                                    # lanjut ke solo_pernrs
                                else:
                                    # Log RETURN RFC
                                    for r in (resp.get("RETURN") or []):
                                        typ = (r.get("TYPE") or "").upper()
                                        msg = r.get("MESSAGE") or ""
                                        (logger.warning if typ in ("E", "A") else logger.info)(
                                            f"[SAP-{typ}] {msg}"
                                        )

                                    t_data = resp.get("T_DATA") or []
                                    sap_rows = len(t_data)

                                    logger.info(
                                        f"[SAP GROUP] rows={sap_rows} "
                                        f"for {group_pernrs}@{arbpl}/{werks} {begda}"
                                    )

                                    # --- Jika SAP 0 row dan ALLOW_EMPTY_DELETE=False -> tidak hapus apa-apa ---
                                    skipped_empty = False
                                    if sap_rows == 0 and not ALLOW_EMPTY_DELETE:
                                        skipped_empty = True
                                        item_pairs_report.append(
                                            {
                                                "arbpl": arbpl,
                                                "werks": werks,
                                                "mode": "group",
                                                "ok": True,
                                                "sap_rows": 0,
                                                "deleted": 0,
                                                "inserted": 0,
                                                "no_change": True,
                                                "skipped_empty": True,
                                                "rows": 0,
                                                "del": 0,
                                                "ins": 0,
                                                "skipped": True,
                                                "pernrs": pernrs_for_rfc,
                                            }
                                        )
                                    else:
                                        # Normal: hapus & insert
                                        try:
                                            deleted = delete_old_rows(
                                                cur, arbpl, werks, begda, endda, pernrs_for_rfc
                                            )
                                            db.commit()
                                            deleted_pair += deleted
                                            logger.info(
                                                f"[DB GROUP] deleted={deleted} "
                                                f"for {pernrs_for_rfc} @ {arbpl}/{werks} {begda}..{endda}"
                                            )
                                        except mysql.connector.Error as e:
                                            db.rollback()
                                            logger.error(f"[DB DELETE GROUP] {e}")
                                            item_pairs_report.append(
                                                {
                                                    "arbpl": arbpl,
                                                    "werks": werks,
                                                    "mode": "group",
                                                    "ok": False,
                                                    "error": f"DB delete: {e}",
                                                    "pernrs": pernrs_for_rfc,
                                                }
                                            )
                                            # lompat ke solo_pernrs, tapi data lama grp tidak aman
                                        else:
                                            # Insert baru
                                            try:
                                                rows: List[Dict[str, Any]] = []
                                                for r in t_data:
                                                    pernr_row = str(r.get("PERNR") or "").strip()
                                                    if pernr_row.isdigit():
                                                        pernr_row = pernr_row.zfill(8)
                                                    role_val = role_map.get(pernr_row)
                                                    rows.append(
                                                        normalize_tdata(
                                                            r,
                                                            arbpl,
                                                            werks,
                                                            desc_val,
                                                            role_val,
                                                            devisi_val,
                                                        )
                                                    )

                                                if rows:
                                                    BATCH = 500
                                                    for i_b in range(0, len(rows), BATCH):
                                                        cur.executemany(
                                                            UPSERT_SQL_YPPR, rows[i_b : i_b + BATCH]
                                                        )
                                                    db.commit()
                                                    inserted = len(rows)
                                                    inserted_pair += inserted
                                                    logger.info(
                                                        f"[DB GROUP] inserted={inserted} "
                                                        f"for {group_pernrs}@{arbpl}/{werks} {begda}"
                                                    )
                                            except mysql.connector.Error as e:
                                                db.rollback()
                                                logger.error(f"[DB INSERT GROUP] {e}")
                                                item_pairs_report.append(
                                                    {
                                                        "arbpl": arbpl,
                                                        "werks": werks,
                                                        "mode": "group",
                                                        "ok": False,
                                                        "error": f"DB insert: {e}",
                                                        "deleted": deleted,
                                                        "sap_rows": sap_rows,
                                                        "pernrs": pernrs_for_rfc,
                                                    }
                                                )
                                            else:
                                                item_pairs_report.append(
                                                    {
                                                        "arbpl": arbpl,
                                                        "werks": werks,
                                                        "mode": "group",
                                                        "ok": True,
                                                        "deleted": deleted,
                                                        "inserted": inserted,
                                                        "sap_rows": sap_rows,
                                                        "skipped_empty": skipped_empty,
                                                        "del": deleted,
                                                        "ins": inserted,
                                                        "rows": sap_rows,
                                                        "skipped": skipped_empty,
                                                        "pernrs": pernrs_for_rfc,
                                                    }
                                                )

                        # ======================================================
                        # 4b) PROSES SOLO (EV_SUBRC == 0) -> TANPA INDUK
                        # ======================================================
                        for solo_pernr in solo_pernrs:
                            # Solo: tidak pernah menambah induk, hanya 1 NIK ini saja
                            pernrs_for_rfc = [solo_pernr]

                            # filter blacklist (mestinya sudah, tapi jaga-jaga)
                            if BLACKLIST_PERNRS and is_blacklisted(solo_pernr):
                                logger.info(
                                    f"[REFRESH SOLO] Skip {solo_pernr} karena BLACKLIST "
                                    f"untuk {arbpl}/{werks}"
                                )
                                continue

                            try:
                                logger.info(
                                    f"[SAP CALL SOLO] {arbpl}/{werks} {begda}..{endda} "
                                    f"T_PERNR={[solo_pernr]}"
                                )
                                resp = call_rfc_yppr(
                                    rfc, arbpl, werks, begda, endda, pernrs_for_rfc
                                )
                            except (ABAPApplicationError, ABAPRuntimeError, CommunicationError) as e:
                                logger.error(f"[SAP SOLO] {e}")
                                item_pairs_report.append(
                                    {
                                        "arbpl": arbpl,
                                        "werks": werks,
                                        "mode": "solo",
                                        "ok": False,
                                        "error": f"SAP: {e}",
                                        "pernrs": pernrs_for_rfc,
                                    }
                                )
                                continue

                            # Log RETURN
                            for r in (resp.get("RETURN") or []):
                                typ = (r.get("TYPE") or "").upper()
                                msg = r.get("MESSAGE") or ""
                                (logger.warning if typ in ("E", "A") else logger.info)(
                                    f"[SAP-{typ}] {msg}"
                                )

                            t_data = resp.get("T_DATA") or []
                            sap_rows = len(t_data)

                            logger.info(
                                f"[SAP SOLO] rows={sap_rows} "
                                f"for {solo_pernr}@{arbpl}/{werks} {begda}"
                            )

                            skipped_empty = False
                            if sap_rows == 0 and not ALLOW_EMPTY_DELETE:
                                skipped_empty = True
                                item_pairs_report.append(
                                    {
                                        "arbpl": arbpl,
                                        "werks": werks,
                                        "mode": "solo",
                                        "ok": True,
                                        "sap_rows": 0,
                                        "deleted": 0,
                                        "inserted": 0,
                                        "no_change": True,
                                        "skipped_empty": True,
                                        "rows": 0,
                                        "del": 0,
                                        "ins": 0,
                                        "skipped": True,
                                        "pernrs": pernrs_for_rfc,
                                    }
                                )
                                continue

                            # Hapus data lama hanya untuk solo_pernr ini
                            try:
                                deleted = delete_old_rows(
                                    cur, arbpl, werks, begda, endda, pernrs_for_rfc
                                )
                                db.commit()
                                deleted_pair += deleted
                                logger.info(
                                    f"[DB SOLO] deleted={deleted} "
                                    f"for {pernrs_for_rfc} @ {arbpl}/{werks} {begda}..{endda}"
                                )
                            except mysql.connector.Error as e:
                                db.rollback()
                                logger.error(f"[DB DELETE SOLO] {e}")
                                item_pairs_report.append(
                                    {
                                        "arbpl": arbpl,
                                        "werks": werks,
                                        "mode": "solo",
                                        "ok": False,
                                        "error": f"DB delete: {e}",
                                        "pernrs": pernrs_for_rfc,
                                    }
                                )
                                continue

                            # Insert baru
                            try:
                                rows: List[Dict[str, Any]] = []
                                for r in t_data:
                                    pernr_row = str(r.get("PERNR") or "").strip()
                                    if pernr_row.isdigit():
                                        pernr_row = pernr_row.zfill(8)
                                    role_val = role_map.get(pernr_row)
                                    rows.append(
                                        normalize_tdata(
                                            r,
                                            arbpl,
                                            werks,
                                            desc_val,
                                            role_val,
                                            devisi_val,
                                        )
                                    )

                                inserted = 0
                                if rows:
                                    BATCH = 500
                                    for i_b in range(0, len(rows), BATCH):
                                        cur.executemany(
                                            UPSERT_SQL_YPPR, rows[i_b : i_b + BATCH]
                                        )
                                    db.commit()
                                    inserted = len(rows)
                                    inserted_pair += inserted
                                    logger.info(
                                        f"[DB SOLO] inserted={inserted} "
                                        f"for {solo_pernr}@{arbpl}/{werks} {begda}"
                                    )
                            except mysql.connector.Error as e:
                                db.rollback()
                                logger.error(f"[DB INSERT SOLO] {e}")
                                item_pairs_report.append(
                                    {
                                        "arbpl": arbpl,
                                        "werks": werks,
                                        "mode": "solo",
                                        "ok": False,
                                        "error": f"DB insert: {e}",
                                        "deleted": deleted,
                                        "sap_rows": sap_rows,
                                        "pernrs": pernrs_for_rfc,
                                    }
                                )
                                continue

                            item_pairs_report.append(
                                {
                                    "arbpl": arbpl,
                                    "werks": werks,
                                    "mode": "solo",
                                    "ok": True,
                                    "deleted": deleted,
                                    "inserted": inserted,
                                    "sap_rows": sap_rows,
                                    "skipped_empty": skipped_empty,
                                    "del": deleted,
                                    "ins": inserted,
                                    "rows": sap_rows,
                                    "skipped": skipped_empty,
                                    "pernrs": pernrs_for_rfc,
                                }
                            )

                    finally:
                        release_pair_lock(cur, arbpl, werks)

                    total_deleted += deleted_pair
                    total_inserted += inserted_pair

                # --- Summary per item (JSON response) ---
                results.append(
                    {
                        "ok": any(p.get("ok") for p in item_pairs_report),
                        "pernr": pernr,
                        "strategy": strategy,
                        "pairs_used": item_pairs_report,
                        "deleted_total": total_deleted,
                        "inserted_total": total_inserted,
                        # field lama vs baru tetap diisi semua biar aman
                        "report": item_pairs_report,
                        "del": total_deleted,
                        "ins": total_inserted,
                    }
                )

            finally:
                # apapun hasilnya, item ke-idx dianggap sudah diproses
                update_progress(
                    job_id,
                    idx,
                    pernr=pernr,
                    begda=begda_for_progress,
                    arbpl=arbpl_for_progress,
                    werks=werks_for_progress,
                )

    except Exception as e:
        logger.error(f"[YPPR058] System Error: {e}")
        finish_progress(job_id, ok=False, error=str(e))
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass
        try:
            rfc.close()
        except Exception:
            pass
        return jsonify({"ok": False, "error": f"System Error: {str(e)}", "job_id": job_id}), 500

    # normal success
    cur.close()
    db.close()
    try:
        rfc.close()
    except Exception:
        pass

    finish_progress(job_id, ok=True)

    return jsonify({"ok": True, "results": results, "job_id": job_id})

# ============================================================================
# ENDPOINT 2: REFRESH WC PERSON (MASTER DATA) - dengan info PERNR baru
# ============================================================================
@app.post("/api/wc_person/sync")
def api_sync_wc_person():
    """
    Payload: {"arbpl": "WC033", "werks": "2000"}

    Logika:
      1. CR_PERSONS_OF_WORKCENTER  -> tarik master WC person
      2. Hapus data lama WC tsb di wc_person_data, lalu insert data baru
      3. Z_RFC_DISPLAY_NIK_CONF    -> update role INDUK
      4. Z_FM_GET_WC_DESC          -> update deskripsi WC
      5. Excel DEVISI.xlsx         -> update kolom devisi
      6. Hitung NIK baru vs data lama, kembalikan:
         - pernrs        : semua NIK aktif (setelah blacklist)
         - pernrs_new    : hanya NIK baru
         - pernrs_new_count, pernrs_old_count
    """
    # --- Ambil & validasi payload ---
    try:
        data = request.get_json(force=True) or {}
        arbpl = str(data.get("arbpl", "")).strip().upper()
        werks = str(data.get("werks", "")).strip()
    except Exception:
        return jsonify({"ok": False, "message": "Invalid JSON"}), 400

    if not arbpl or not werks:
        return jsonify({"ok": False, "message": "Harap isi Work Center dan Plant"}), 400

    logger.info(f"[WC SYNC] START REQUEST: {arbpl} plant {werks}")

    conn: Optional[Connection] = None
    db = None

    try:
        # --- Koneksi ke SAP & MySQL ---
        sap_params = {**DEFAULT_SAP, "user": SAP_USERNAME, "passwd": SAP_PASSWORD}
        conn = Connection(**sap_params)

        db = get_mysql()
        cur = db.cursor()

        # (Opsional) Hitung jumlah data lama, hanya untuk log
        try:
            cur.execute(COUNT_WC_SQL, (arbpl, werks))
            old_count_row = cur.fetchone()
            old_count = int(old_count_row[0]) if old_count_row else 0
        except Exception:
            old_count = 0

        # NEW: ambil daftar PERNR lama (sebelum dihapus) untuk deteksi NIK baru
        try:
            cur.execute(GET_OLD_PERNRS_SQL, (arbpl, werks))
            old_pernrs_set = {
                _pad8(str(p)) for (p,) in cur.fetchall() if p
            }
        except Exception:
            old_pernrs_set = set()

        # ------------------------------------------------------------------
        # 1. TARIK DATA UTAMA dari CR_PERSONS_OF_WORKCENTER
        # ------------------------------------------------------------------
        try:
            res = conn.call(RFC_MAIN_WC, ARBPL=arbpl, WERKS=werks, DATE="99991231")
        except Exception as e:
            logger.error(f"[WC SYNC] RFC Error {RFC_MAIN_WC}: {e}")
            return jsonify(
                {"ok": False, "message": f"RFC Error ({RFC_MAIN_WC}): {str(e)}"}
            ), 500

        # Cari tabel list pertama yang berisi data
        raw_rows: List[Dict[str, Any]] = []
        for _, v in res.items():
            if isinstance(v, list) and v:
                raw_rows = v
                break

        # Jika SAP mengembalikan kosong -> hapus semua WC lama & selesai
        if not raw_rows:
            # Hapus master WC
            cur.execute(DELETE_WC_SQL, (arbpl, werks))
            deleted_rows = cur.rowcount

            # Hapus semua detail yppr058_data untuk WC/Plant ini
            yppr_deleted = 0
            try:
                cur.execute(DELETE_YPPR_WC_SQL, (arbpl, werks))
                yppr_deleted = cur.rowcount
                db.commit()
            except Exception as e:
                db.rollback()
                logger.error(
                    f"[WC SYNC] Gagal hapus {OUT_TABLE} untuk {arbpl}/{werks}: {e}"
                )

            logger.info(
                f"[WC SYNC] {arbpl} kosong di SAP. "
                f"Old={old_count}, DeletedMaster={deleted_rows}, "
                f"DeletedDetail={yppr_deleted}"
            )
            cur.close()
            return jsonify(
                {
                    "ok": True,
                    "message": f"WC {arbpl} kosong di SAP. {deleted_rows} data lama dihapus.",
                    "deleted": deleted_rows,
                    "inserted": 0,
                    "desc": "",
                    "role_induk_count": 0,
                    "devisi_updated": False,
                    "pernrs": [],
                    "pernrs_count": 0,
                    # NEW: info NIK lama vs baru
                    "pernrs_new": [],
                    "pernrs_new_count": 0,
                    "pernrs_old_count": len(old_pernrs_set),
                    # semua NIK lama dianggap removed
                    "pernrs_removed": sorted(old_pernrs_set),
                    "pernrs_removed_count": len(old_pernrs_set),
                    "yppr058_deleted": yppr_deleted,
                }
            )

        # Normalisasi baris SAP ke struktur wc_person_data
        norm_rows: List[Dict[str, Any]] = [
            normalize_wc_row(r, arbpl, werks) for r in raw_rows
        ]

        # Terapkan blacklist PERNR (skip dari insert & dari daftar pernrs)
        if BLACKLIST_PERNRS:
            before = len(norm_rows)
            norm_rows = [r for r in norm_rows if not is_blacklisted(r.get("pernr", ""))]
            skipped = before - len(norm_rows)
            if skipped:
                logger.info(
                    f"[WC SYNC] BLACKLIST: {skipped} baris di-skip (PERNR) untuk {arbpl}/{werks}"
                )

        # Daftar PERNR unik (LIST) untuk dikirim ke frontend & dipakai cek role
        pernrs: List[str] = sorted(
            {r["pernr"] for r in norm_rows if r.get("pernr")}
        )

        # NEW: hitung NIK baru dibanding data lama di DB
        new_pernrs_set = set(pernrs) - old_pernrs_set
        pernrs_new: List[str] = sorted(new_pernrs_set)

        removed_pernrs_set = old_pernrs_set - set(pernrs)
        pernrs_removed: List[str] = sorted(removed_pernrs_set)

        yppr058_deleted = 0
        if pernrs_removed:
            placeholders = ",".join(["%s"] * len(pernrs_removed))
            sql_del_yppr = (
                f"DELETE FROM `{OUT_TABLE}` "
                f"WHERE `arbpl`=%s AND `werks`=%s AND `pernr` IN ({placeholders})"
            )
            params = [arbpl, werks] + pernrs_removed
            try:
                cur.execute(sql_del_yppr, params)
                yppr058_deleted = cur.rowcount
                db.commit()
                logger.info(
                    f"[WC SYNC] Deleted {yppr058_deleted} rows from {OUT_TABLE} "
                    f"for removed PERNRs {pernrs_removed} at {arbpl}/{werks}"
                )
            except Exception as e:
                db.rollback()
                logger.error(
                    f"[WC SYNC] Gagal hapus {OUT_TABLE} untuk NIK {pernrs_removed} "
                    f"{arbpl}/{werks}: {e}"
                )
        # ------------------------------------------------------------------
        # 2. DELETE data lama WC ini -> INSERT data baru
        # ------------------------------------------------------------------
        cur.execute(DELETE_WC_SQL, (arbpl, werks))
        deleted_rows = cur.rowcount  # jumlah yang benar-benar terhapus

        BATCH = 500
        for i in range(0, len(norm_rows), BATCH):
            cur.executemany(UPSERT_WC_SQL, norm_rows[i : i + BATCH])
        db.commit()

        inserted_rows = len(norm_rows)
        logger.info(
            f"[WC SYNC] {arbpl} Update DB. Old={old_count}, Deleted={deleted_rows}, Inserted={inserted_rows}"
        )

        # ------------------------------------------------------------------
        # 3. CEK ROLE INDUK (Z_RFC_DISPLAY_NIK_CONF)
        # ------------------------------------------------------------------
        pernrs_set = set(pernrs)
        active_induk: set[str] = set()

        for p in pernrs_set:
            try:
                r_role = conn.call(RFC_ROLE_WC, PERNR=p)
                tab = r_role.get("RESULT_TAB") or []
                # Kalau ada baris yg DELETED != 'X' -> masih aktif sebagai induk
                if any(str(row.get("DELETED", "")).strip().upper() != "X" for row in tab):
                    active_induk.add(p)
            except Exception as e:
                logger.warning(f"[WC SYNC] Role check failed for {p}: {e}")

        if active_induk:
            role_data = [("INDUK", p) for p in active_induk]
            cur.executemany(UPDATE_ROLE_WC_SQL, role_data)
            db.commit()
            logger.info(
                f"[WC SYNC] Updated Role INDUK for {len(active_induk)} personil."
            )

        # ------------------------------------------------------------------
        # 4. DESKRIPSI WC (Z_FM_GET_WC_DESC)
        # ------------------------------------------------------------------
        wc_desc = ""
        try:
            r_desc = conn.call(RFC_DESC_WC, IV_ARBPL=arbpl, IV_WERKS=werks)
            wc_desc = (r_desc.get("E_DESC") or "").strip()
        except Exception as e:
            logger.warning(f"[WC SYNC] Gagal ambil deskripsi WC: {e}")

        if wc_desc:
            cur.execute(UPDATE_DESC_WC_SQL, (wc_desc, arbpl, werks))
            db.commit()
            logger.info(f"[WC SYNC] Updated Description: {wc_desc}")

        # ------------------------------------------------------------------
        # 5. UPDATE DEVISI & KODE LARAVEL dari Excel DEVISI.xlsx
        # ------------------------------------------------------------------
        devisi_updated = False
        kolar_updated_val = "" 

        try:
            devisi_mapping = load_devisi_mapping_from_excel(DEVISI_FILE)
            key = (arbpl, werks)
            
            if key in devisi_mapping:
                data_excel = devisi_mapping[key]

                # Devisi boleh kosong -> simpan sebagai "" saja
                devisi_val = data_excel.get("devisi") or ""

                # Kode Laravel: kalau kosong di Excel -> None (NULL di DB)
                raw_kolar = (data_excel.get("kode_laravel") or "").strip()
                kolar_val = raw_kolar if raw_kolar else None

                # Jalankan Update SQL Baru
                cur.execute(UPDATE_EXTRA_INFO_SQL, (devisi_val, kolar_val, arbpl, werks))
                db.commit()

                devisi_updated = True
                kolar_updated_val = kolar_val or ""
                logger.info(
                    f"[WC SYNC] Updated Excel Info for {arbpl}: "
                    f"Devisi='{devisi_val}', Laravel='{kolar_val}'"
                )
            else:
                logger.info(
                    f"[WC SYNC] No Excel mapping found for {arbpl}/{werks}."
                )
        except Exception as e:
            logger.warning(f"[WC SYNC] Gagal update info Excel: {e}")

        cur.close()

        # ------------------------------------------------------------------
        # 6. RESPONSE: kirim juga daftar PERNR lama & baru ke frontend
        # ------------------------------------------------------------------
        total_pernrs_sap = len(
            {normalize_wc_row(r, arbpl, werks)["pernr"] for r in raw_rows if r.get("PERNR")}
        )
        blacklist_skip_count = max(0, total_pernrs_sap - len(pernrs))

        return jsonify(
            {
                "ok": True,
                "message": (
                    f"Sukses! Dihapus: {deleted_rows}, "
                    f"Dimasukkan: {inserted_rows} personil."
                ),
                "desc": wc_desc,
                "role_induk_count": len(active_induk),
                "deleted": deleted_rows,
                "inserted": inserted_rows,
                "devisi_updated": devisi_updated,
                # semua PERNR hasil sync (setelah blacklist)
                "pernrs": pernrs,
                "pernrs_count": len(pernrs),
                # hanya PERNR yang benar-benar baru dibanding data lama
                "pernrs_new": pernrs_new,
                "pernrs_new_count": len(pernrs_new),
                "pernrs_old_count": len(old_pernrs_set),
                "blacklist_skip_count": blacklist_skip_count,
                # NEW: NIK yang hilang & jumlah baris detail yang dihapus
                "pernrs_removed": pernrs_removed,
                "pernrs_removed_count": len(pernrs_removed),
                "yppr058_deleted": yppr058_deleted,
            }
        )
    except Exception as e:
        logger.error(f"[WC SYNC] System Error: {e}")
        if db:
            try:
                db.rollback()
            except Exception:
                pass
        return jsonify({"ok": False, "message": f"System Error: {str(e)}"}), 500
    finally:
        if db:
            try:
                db.close()
            except Exception:
                pass
        if conn:
            try:
                conn.close()
            except Exception:
                pass

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5010, debug=False)
